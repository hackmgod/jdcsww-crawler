#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生产环境增强爬虫 - 完整版
功能:
1. 爬取所有批次(1-403)的新能源车辆完整数据（34种车辆类型）
2. 包含83个完整字段（基本信息+三电参数+图片）
3. 支持增量保存和断点续传
4. 导出为Excel格式
5. 下载车辆图片

配置说明:
- 车辆类型: 34种新能源汽车子类型（纯电动、混合动力、插电式、增程式、换电式等）
- 燃料类型: 留空（不限制）
- 批次范围: 1-403（全部批次）
"""

import urllib.request
import urllib.parse
import json
import time
import re
import os
import logging
import random
from datetime import datetime
from typing import List, Dict, Optional, Set
from http.cookiejar import CookieJar


class EnhancedProductionCrawler:
    """生产环境增强爬虫 - 安全版"""

    # ✅ 有序字段列表（按照Section顺序排列，用于Excel导出）
    ORDERED_FIELDS = [
        # ==================== 1. 列表页字段 ====================
        '批次',
        '燃料类型',
        '车辆名称',
        '详情链接',
        '公告编号',
        '公告型号',
        '生产企业',

        # ==================== 2. 生产企业信息 ====================
        '生产企业信息_车辆名称',
        '生产企业信息_车辆类型',
        '生产企业信息_制造国',
        '生产企业信息_牌照类型',
        '生产企业信息_公告批次',
        '生产企业信息_发布日期',
        '生产企业信息_产品号',
        '生产企业信息_目录序号',
        '生产企业信息_中文品牌',
        '生产企业信息_英文品牌',
        '生产企业信息_公告型号',
        '生产企业信息_车型',
        '生产企业信息_企业名称',
        '生产企业信息_产品ID号',
        '生产企业信息_生产地址',
        '生产企业信息_注册地址',
        '生产企业信息_通讯地址',
        '生产企业信息_电话',
        '生产企业信息_法人代表',
        '生产企业信息_邮编',
        '生产企业信息_传真',
        '生产企业信息_电子邮箱',
        '生产企业信息_联系人',

        # ==================== 3. 主要技术参数 ====================
        '主要技术参数_车辆总质量',
        '主要技术参数_车辆转向形式',
        '主要技术参数_额定载质量',
        '主要技术参数_车辆总轴数',
        '主要技术参数_车辆整备质量',
        '主要技术参数_最高车速',
        '主要技术参数_车辆轴荷',
        '主要技术参数_额定载客',
        '主要技术参数_车辆轴距',
        '主要技术参数_前排允许载客',
        '主要技术参数_车辆前轮距',
        '主要技术参数_后排允许载客',
        '主要技术参数_车辆后轮距',
        '主要技术参数_车辆传动形式',
        '主要技术参数_钢板弹簧片数',
        '主要技术参数_车辆轮胎数',
        '主要技术参数_半挂车鞍座最大载质量',
        '主要技术参数_准拖挂车总质量',
        '主要技术参数_接近角/离去角',
        '主要技术参数_载质量利用系数',
        '主要技术参数_车辆前悬/后悬',
        '主要技术参数_车辆轮胎规格型号',
        '主要技术参数_外形尺寸长Ⅹ宽Ⅹ高',
        '主要技术参数_货箱栏板长Ⅹ宽Ⅹ高',
        '主要技术参数_车辆识别代码',

        # ==================== 4. 车辆燃料参数 ====================
        '车辆燃料参数_车辆燃料种类',
        '车辆燃料参数_车辆耗油',  # ✅ 去除单位 - 解析时会去除
        '车辆燃料参数_排放依据标准',
        '车辆燃料参数_排放标准提示',
        '车辆燃料参数_机动车环保网',
        '车辆燃料参数_是否新能源汽车',
        '车辆燃料参数_新能源汽车种类',

        # ==================== 5. 电芯参数 ====================
        '电芯参数_动力类型',
        '电芯参数_续驶里程',  # ✅ 去除单位和特殊标记
        '电芯参数_电芯类型',
        '电芯参数_电芯型号',
        '电芯参数_电芯形状',
        '电芯参数_电芯尺寸',
        '电芯参数_电芯电压',
        '电芯参数_电芯容量',
        '电芯参数_电芯电量',
        '电芯参数_电芯数量',
        '电芯参数_电芯生产企业',

        # ==================== 6. 电池信息 ====================
        '电池信息_电池容量',
        '电池信息_电池电压',
        '电池信息_电池电量',
        '电池信息_电池质量',
        '电池信息_电池管理系统BMS',
        '电池信息_电池生产企业',
        '电池信息_电池组合方式',

        # ==================== 7. 电机信息 ====================
        '电机信息_电机型号',
        '电机信息_电机类型',
        '电机信息_电机冷却方式',
        '电机信息_电机数量',
        '电机信息_电机峰值功率/转速/转矩',
        '电机信息_电机生产企业',

        # ==================== 8. 电控系统 ====================
        '电控系统_电控供应商',
        '电控系统_充电机型号',
        '电控系统_充电机生产企业',
        '电控系统_充电机功率',
        '电控系统_车载DC/DC变换器',
        '电控系统_高压配电盒',

        # ==================== 9. 其他 ====================
        '其他',

        # ==================== 10. 其他字段 ====================
        '详情页URL',
        '爬取时间',
        '车辆图片',
    ]

    def __init__(self):
        self.base_url = "https://www.jdcsww.com"
        self.cookie_jar = CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar)
        )
        self.output_dir = 'data/production_enhanced'
        self.image_dir = os.path.join(self.output_dir, 'images')
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.image_dir, exist_ok=True)

        # 状态跟踪
        self.state_file = os.path.join(self.output_dir, 'crawler_state.json')
        self.data_file = os.path.join(self.output_dir, 'all_vehicles_complete.json')
        self.excel_file = os.path.join(self.output_dir, 'vehicles_complete_all.xlsx')

        # ⚡ 反爬虫延迟配置（超保守版 - 2026-04-02优化）
        self.detail_delay_min = 45.0      # 详情页之间最小延迟（秒）- 从10秒提升到45秒
        self.detail_delay_max = 90.0      # 详情页之间最大延迟（秒）- 从20秒提升到90秒
        self.batch_delay_min = 120.0      # 批次间最小延迟（秒）- 从15秒提升到2分钟
        self.batch_delay_max = 300.0      # 批次间最大延迟（秒）- 从30秒提升到5分钟
        self.rest_interval = 3            # 每N个批次休息 - 从20改为3（更频繁休息）
        self.rest_duration = 900          # 休息时长（秒）- 从60秒提升到15分钟

        # 统计信息
        self.request_count = 0
        self.blocked_count = 0

        # ✅ 延迟等级配置（2026-04-03添加）
        self.delay_profiles = {
            '1': {  # 🔥 极速
                'name': '极速',
                'emoji': '🔥',
                'detail_min': 0.5,
                'detail_max': 2.0,
                'batch_min': 5.0,
                'batch_max': 10.0,
                'rest_interval': 10,
                'rest_duration': 60,
                'risk': '极高风险',
                'description': '仅用于代码验证（1-5辆车）'
            },
            '2': {  # ⚡ 快速
                'name': '快速',
                'emoji': '⚡',
                'detail_min': 3.0,
                'detail_max': 8.0,
                'batch_min': 15.0,
                'batch_max': 30.0,
                'rest_interval': 5,
                'rest_duration': 120,
                'risk': '高风险',
                'description': '测试模式，不频繁测试不会被检测'
            },
            '3': {  # ⚖️ 标准
                'name': '标准',
                'emoji': '⚖️',
                'detail_min': 20.0,
                'detail_max': 40.0,
                'batch_min': 60.0,
                'batch_max': 120.0,
                'rest_interval': 3,
                'rest_duration': 300,
                'risk': '中等风险',
                'description': '小范围爬取（10-50个批次）'
            },
            '4': {  # 🛡️ 安全
                'name': '安全',
                'emoji': '🛡️',
                'detail_min': 45.0,
                'detail_max': 90.0,
                'batch_min': 120.0,
                'batch_max': 300.0,
                'rest_interval': 3,
                'rest_duration': 900,
                'risk': '较安全',
                'description': '正常爬取（推荐）'
            },
            '5': {  # 🔒 超保守
                'name': '超保守',
                'emoji': '🔒',
                'detail_min': 90.0,
                'detail_max': 180.0,
                'batch_min': 300.0,
                'batch_max': 600.0,
                'rest_interval': 2,
                'rest_duration': 1800,
                'risk': '最安全',
                'description': '全量爬取403批次'
            }
        }

        # 配置日志
        self._setup_logger()

    def _setup_logger(self):
        """配置日志系统"""
        self.logger = logging.getLogger('EnhancedCrawler')
        self.logger.setLevel(logging.INFO)

        if self.logger.handlers:
            self.logger.handlers.clear()

        formatter = logging.Formatter(
            '%(asctime)s | %(levelname)-8s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

        log_file = os.path.join(self.output_dir, f'crawler_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)

        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)

        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        self.log_file = log_file

        self.logger.info("=" * 70)
        self.logger.info("增强版爬虫启动")
        self.logger.info(f"日志文件: {log_file}")
        self.logger.info("=" * 70)

    def is_working_hours(self) -> bool:
        """判断是否在工作时间内（9:00-18:00）"""
        current_hour = datetime.now().hour
        is_working = 9 <= current_hour < 18
        if not is_working:
            self.logger.info(f"⏰ 当前时间{datetime.now().strftime('%H:%M')}，不在工作时间内")
        return is_working

    def ask_working_hours(self) -> bool:
        """询问是否启用工作时间限制"""
        print("\n" + "⏰" * 35)
        print("工作时间限制设置")
        print("⏰" * 35)
        print("\n💡 启用工作时间限制:")
        print("  ✓ 仅在 09:00-18:00 之间运行爬虫")
        print("  ✓ 其他时间自动退出并保存进度")
        print("  ✓ 模拟正常用户工作习惯，降低被封风险")
        print("  ✓ 适合长期稳定的爬取任务")
        print("\n💡 禁用工作时间限制:")
        print("  ✓ 24小时不间断运行")
        print("  ✓ 适合快速完成测试任务")
        print("  ✓ 风险略高（夜间频繁请求可能被检测）")
        print("\n" + "⏰" * 35)

        work_hours = input("\n是否启用工作时间限制? [y/n]: ").strip().lower()
        enable = work_hours == 'y'

        if enable:
            print("\n✅ 工作时间限制已启用")
            print("   ⏰ 运行时间: 09:00-18:00")
            print("   ⏰ 休息时间: 18:00-次日09:00")
            print("   💡 非工作时间会自动退出并保存进度")
        else:
            print("\n⚠️  工作时间限制已禁用")
            print("   ⏰ 运行时间: 24小时")
            print("   ⚠️  注意: 夜间频繁请求可能增加被封风险")

        print()
        return enable

    def ask_export_excel(self) -> bool:
        """询问是否自动导出Excel"""
        print("\n" + "📊" * 35)
        print("Excel导出设置")
        print("📊" * 35)
        print("\n💡 爬取完成后是否自动导出Excel:")
        print("  ✓ 导出: 爬取完成自动生成Excel文件")
        print("  ✓ 导出: 适合最终数据整理和查看")
        print("  ✗ 不导出: 仅保存JSON格式，速度更快")
        print("  ✗ 不导出: 可以后续手动导出（选项5）")
        print("\n" + "📊" * 35)

        export_excel = input("\n爬取完成后是否自动导出Excel? [y/n]: ").strip().lower()
        enable = export_excel == 'y'

        if enable:
            print("\n✅ 已启用: 爬取完成后自动导出Excel")
            print(f"   📄 文件路径: {self.excel_file}")
        else:
            print("\n⚠️  已禁用: 爬取完成后不自动导出Excel")
            print("   💡 后续可使用选项5手动导出")

        print()
        return enable

    def select_delay_profile(self, include_custom=True, exclude_fastest=False):
        """
        让用户选择延迟配置

        Args:
            include_custom: 是否包含自定义选项
            exclude_fastest: 是否排除极速选项（仅用于安全考虑）

        Returns:
            profile_key: 延迟等级的key（'1'-'5' 或 'custom'）
        """
        print("\n" + "⏱️ " * 35)
        print("延迟配置选择")
        print("⏱️ " * 35)
        print("\n请选择延迟配置：\n")

        # 显示所有延迟等级
        for key, profile in self.delay_profiles.items():
            # 排除极速选项（如果要求）
            if exclude_fastest and key == '1':
                continue

            emoji = profile['emoji']
            name = profile['name']
            detail_range = f"{profile['detail_min']}-{profile['detail_max']}"
            batch_range = f"{profile['batch_min']}-{profile['batch_max']}"
            risk = profile['risk']
            desc = profile['description']

            print(f"{key}. {emoji} {name} - 详情{detail_range}秒 / 批次{batch_range}秒（{risk}）")
            print(f"   → {desc}")

        # 显示自定义选项
        if include_custom:
            custom_key = str(len(self.delay_profiles) + 1)
            print(f"{custom_key}. ✏️  自定义 - 手动输入所有参数")

        print("\n" + "⏱️ " * 35)

        # 获取用户选择
        max_choice = str(len(self.delay_profiles) + (1 if include_custom else 0))
        choice = input(f"\n请选择延迟等级 (1-{max_choice}): ").strip()

        # 验证输入
        if choice not in self.delay_profiles and (not include_custom or choice != custom_key):
            print("⚠️  输入无效，使用默认配置：安全（🛡️）")
            choice = '4'  # 默认安全等级

        # 如果选择自定义
        if include_custom and choice == custom_key:
            return 'custom'

        return choice

    def custom_delay_config(self):
        """让用户自定义延迟配置"""
        print("\n" + "="*70)
        print("✏️  自定义延迟配置")
        print("="*70)
        print("\n【详情页延迟】（爬取一辆车后，等待多久再爬下一辆）")

        try:
            detail_min = float(input("  最小延迟（秒）: ").strip())
            detail_max = float(input("  最大延迟（秒）: ").strip())

            if detail_min >= detail_max:
                print("⚠️  最大延迟必须大于最小延迟，已自动调整")
                detail_max = detail_min + 5

            print("\n【批次间延迟】（一个批次完成后，等待多久开始下一批次）")
            batch_min = float(input("  最小延迟（秒）: ").strip())
            batch_max = float(input("  最大延迟（秒）: ").strip())

            if batch_min >= batch_max:
                print("⚠️  最大延迟必须大于最小延迟，已自动调整")
                batch_max = batch_min + 30

            print("\n【休息策略】（可选，输入0则不休息）")
            rest_input = input("  每几个批次休息一次（0=不休息）: ").strip()
            rest_interval = int(rest_input) if rest_input else 0

            rest_duration = 0
            if rest_interval > 0:
                rest_input = input("  每次休息时长（秒）: ").strip()
                rest_duration = int(rest_input) if rest_input else 300

            # 构建自定义配置
            custom_profile = {
                'name': '自定义',
                'emoji': '✏️',
                'detail_min': detail_min,
                'detail_max': detail_max,
                'batch_min': batch_min,
                'batch_max': batch_max,
                'rest_interval': rest_interval if rest_interval > 0 else 999,  # 0改为999（不休息）
                'rest_duration': rest_duration if rest_interval > 0 else 0,
                'risk': '自定义',
                'description': '用户自定义延迟'
            }

            print("\n" + "="*70)
            print("📊 自定义配置确认")
            print("="*70)
            print(f"详情页延迟: {detail_min}-{detail_max}秒")
            print(f"批次间延迟: {batch_min}-{batch_max}秒")
            if rest_interval > 0:
                print(f"休息策略: 每{rest_interval}批次休息{rest_duration}秒")
            else:
                print(f"休息策略: 不休息")
            print("="*70)

            confirm = input("\n确认配置？(yes/no): ").strip().lower()
            if confirm == 'yes':
                return custom_profile
            else:
                print("⚠️  已取消自定义，使用默认配置")
                return None

        except ValueError:
            print("❌ 输入无效，使用默认配置")
            return None

    def apply_delay_profile(self, profile_key_or_profile):
        """
        应用延迟配置

        Args:
            profile_key_or_profile: 延迟等级key（'1'-'5'）或自定义配置字典
        """
        profile = None

        # 如果传入的是key，从delay_profiles中获取
        if isinstance(profile_key_or_profile, str):
            profile = self.delay_profiles.get(profile_key_or_profile)
            if not profile:
                print("⚠️  无效的延迟等级，使用默认配置")
                profile = self.delay_profiles['4']
        else:
            # 如果传入的是自定义配置字典
            profile = profile_key_or_profile

        # 应用配置
        self.detail_delay_min = profile['detail_min']
        self.detail_delay_max = profile['detail_max']
        self.batch_delay_min = profile['batch_min']
        self.batch_delay_max = profile['batch_max']
        self.rest_interval = profile['rest_interval']
        self.rest_duration = profile['rest_duration']

        return profile

    def estimate_time(self, num_batches, num_vehicles, profile, include_detail=True):
        """
        估算爬取时间

        Args:
            num_batches: 批次数量
            num_vehicles: 预计车辆数量
            profile: 延迟配置
            include_detail: 是否包含详情页爬取

        Returns:
            估算时间字符串
        """
        if not include_detail:
            # 只爬列表页，每个批次约2秒
            total_seconds = num_batches * 2 * 2  # 纯电动 + 混合动力
            minutes = total_seconds / 60
            if minutes < 60:
                return f"约{int(minutes)}分钟"
            else:
                hours = minutes / 60
                return f"约{int(hours)}小时"

        # 爬取详情页的时间估算
        avg_detail_delay = (profile['detail_min'] + profile['detail_max']) / 2
        detail_time = num_vehicles * avg_detail_delay

        # 批次间延迟估算
        if num_batches > 1:
            avg_batch_delay = (profile['batch_min'] + profile['batch_max']) / 2
            batch_time = (num_batches - 1) * avg_batch_delay
        else:
            batch_time = 0

        # 休息时间估算
        if num_batches > profile['rest_interval'] and profile['rest_interval'] < 100:
            rest_count = num_batches // profile['rest_interval']
            rest_time = rest_count * profile['rest_duration']
        else:
            rest_time = 0

        # 总时间
        total_seconds = detail_time + batch_time + rest_time

        # 转换为可读格式
        minutes = total_seconds / 60
        hours = minutes / 60
        days = hours / 24

        if days >= 1:
            return f"约{days:.1f}天（{int(hours)}小时）"
        elif hours >= 1:
            return f"约{int(hours)}小时（{int(minutes)}分钟）"
        else:
            return f"约{int(minutes)}分钟"

    def _get_chrome_headers(self) -> Dict[str, str]:
        """生成真实Chrome浏览器的完整请求头"""
        return {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'identity',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
            'Sec-Ch-Ua': '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"macOS"',
            'Cache-Control': 'max-age=0',
        }

    def warm_up(self):
        """预热机制 - 仅访问首页，避免触发反爬"""
        self.logger.info("=" * 70)
        self.logger.info("🔥 预热中 - 模拟真实用户访问模式")
        self.logger.info("=" * 70)

        try:
            # 1. 访问首页获取初始Cookie
            self.logger.info("📍 [1/2] 访问首页获取Cookie...")
            homepage = self.fetch_with_session(self.base_url, referer=None, max_retries=2)
            if homepage:
                warm_delay = random.uniform(3, 6)
                self.logger.info(f"✓ 首页访问成功，等待 {warm_delay:.1f} 秒...")
                time.sleep(warm_delay)
            else:
                self.logger.warning("✗ 首页访问失败，继续尝试")

            # 2. 最终停留（不访问空白查询页，避免触发封禁）
            final_delay = random.uniform(2, 5)
            self.logger.info(f"📍 [2/2] 预热完成，最终等待 {final_delay:.1f} 秒...")
            time.sleep(final_delay)

            self.logger.info("✅ 预热完成！开始爬取...")
            self.logger.info("💡 跳过空白查询页，直接访问带参数的URL")
            self.logger.info("=" * 70)
            print()

        except Exception as e:
            self.logger.warning(f"⚠️ 预热过程出现异常: {e}")
            self.logger.info("继续尝试爬取...")

    def is_blocked(self, html: str) -> bool:
        """检测是否被封禁/拦截"""
        if not html:
            return True

        block_indicators = [
            '非正常访问',
            '爬取数据嫌疑',
            '访问过于频繁',
            '您的访问已被限制',
            '请稍后再试',
            '系统检测到异常',
            'Anti-Spam',
            '验证码',
            'captcha',
        ]

        for indicator in block_indicators:
            if indicator in html:
                self.logger.warning(f"🚫 检测到封禁指示: '{indicator}'")
                return True

        return False

    def load_state(self) -> Dict:
        """加载爬虫状态 - 兼容新旧两种失败记录格式"""
        if os.path.exists(self.state_file):
            with open(self.state_file, 'r', encoding='utf-8') as f:
                state = json.load(f)

            # ✅ 兼容旧的失败记录格式（字符串）和新格式（字典）
            if 'failed_batches' in state:
                normalized_failed = []
                for item in state['failed_batches']:
                    if isinstance(item, str):
                        # 旧格式：字符串task_key，转换为新格式
                        normalized_failed.append({
                            'task_key': item,
                            'batch': int(item.split('-')[0]) if '-' in item else 0,
                            'vehicle_type': item.split('-')[1] if '-' in item else '',
                            'vehicle_type_name': '',
                            'fail_time': '',
                            'fail_reason': '旧格式记录',
                            'error_type': 'LegacyFormat'
                        })
                    elif isinstance(item, dict):
                        # 新格式：字典，直接使用
                        normalized_failed.append(item)
                    else:
                        # 未知格式，跳过
                        pass
                state['failed_batches'] = normalized_failed

            return state
        return {
            'completed_batches': [],
            'failed_batches': [],
            'last_update': None,
            'total_vehicles': 0
        }

    def save_state(self, state: Dict):
        """保存爬虫状态"""
        state['last_update'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(self.state_file, 'w', encoding='utf-8') as f:
            json.dump(state, f, ensure_ascii=False, indent=2)

    def load_data(self) -> List[Dict]:
        """加载已有数据 - 自动迁移旧字段名到新字段名"""
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r', encoding='utf-8') as f:
                vehicles = json.load(f)

            # ✅ 字段迁移：将旧字段名转换为新字段名（2026-04-03）

            # 1. 前缀迁移
            field_mapping = {
                '车辆参数_': '主要技术参数_',
                '电芯信息_': '电芯参数_',
            }

            # 2. 单位迁移：去除字段名中的单位（括号内容）
            # ⚠️ 特殊字段白名单：这些字段需要保留单位（已废弃，现在全部去除单位）
            keep_unit_fields = set()  # 空集合 - 所有字段都去除单位

            migrated_count = 0
            for vehicle in vehicles:
                # 迁移1: 前缀替换
                for old_prefix, new_prefix in field_mapping.items():
                    # 查找所有使用旧前缀的字段
                    old_fields = [k for k in vehicle.keys() if k.startswith(old_prefix)]
                    if old_fields:
                        for old_field in old_fields:
                            # 生成新字段名
                            new_field = old_field.replace(old_prefix, new_prefix, 1)
                            # 迁移值到新字段
                            if new_field not in vehicle:
                                vehicle[new_field] = vehicle[old_field]
                            # 删除旧字段
                            del vehicle[old_field]
                            migrated_count += 1

                # 迁移2: 去除字段名中的单位（除了白名单中的字段）
                fields_to_migrate = list(vehicle.keys())
                for field in fields_to_migrate:
                    # 检查是否在白名单中（模糊匹配）
                    should_keep_unit = False
                    for keep_field in keep_unit_fields:
                        # 移除单位后比较
                        keep_field_no_unit = re.sub(r'\([^)]+\)', '', keep_field).strip()
                        field_no_unit = re.sub(r'\([^)]+\)', '', field).strip()
                        if keep_field_no_unit == field_no_unit or keep_field in field:
                            should_keep_unit = True
                            break

                    if should_keep_unit:
                        continue

                    # 提取字段的基础名称（去掉前缀）
                    parts = field.split('_', 1)
                    if len(parts) < 2:
                        continue

                    prefix = parts[0] + '_'
                    base_name = parts[1]

                    # 去除括号中的单位 - 支持中英文括号
                    new_base_name = re.sub(r'[\(（][^)\）]*[\)）]', '', base_name).strip()

                    # ✅ 去除字段名后面的单位标识（KM、mm、Kg、Kw、Ah、V、Wh等）
                    # ⚠️ 注意：不要包含"人"、"个"等，因为它们可能是字段名的一部分（如"联系人"）
                    new_base_name = re.sub(r'(KM|km|MM|mm|KG|kg|KW|kw|AH|ah|V|v|WH|wh|％|%|／|/)$', '', new_base_name).strip()

                    if new_base_name != base_name:
                        new_field = prefix + new_base_name

                        # 迁移值到新字段
                        if new_field not in vehicle:
                            vehicle[new_field] = vehicle[field]
                            migrated_count += 1

                        # 删除旧字段
                        if new_field != field:
                            del vehicle[field]

            if migrated_count > 0:
                self.logger.info(f"✅ 已迁移 {migrated_count} 个旧字段到新字段名")
                # 迁移后自动保存，确保下次不需要再迁移
                self.save_data(vehicles)
                self.logger.info(f"✅ 迁移后的数据已保存")

            return vehicles
        return []

    def save_data(self, vehicles: List[Dict]):
        """保存所有数据"""
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(vehicles, f, ensure_ascii=False, indent=2)

    def fetch_with_session(self, url: str, referer: str = None, max_retries: int = 3) -> Optional[str]:
        """使用会话获取页面 - 增强版（完整请求头+封禁检测）"""
        for attempt in range(max_retries):
            try:
                # 使用完整的Chrome请求头
                headers = self._get_chrome_headers()

                if referer:
                    headers['Referer'] = referer

                req = urllib.request.Request(url, headers=headers)
                with self.opener.open(req, timeout=30) as response:
                    html = response.read().decode('utf-8', errors='ignore')

                    # 检测是否被封禁
                    if self.is_blocked(html):
                        self.blocked_count += 1
                        if attempt < max_retries - 1:
                            retry_delay = 20 + (attempt * 10)
                            self.logger.warning(f"🚫 被拦截，{retry_delay}秒后重试 ({attempt + 1}/{max_retries})")
                            time.sleep(retry_delay)
                            continue
                        else:
                            self.logger.error("❌ 多次重试失败，IP可能已被封禁")
                            self.logger.error("💡 建议：等待15-30分钟后重试，或更换网络")
                            return None

                    self.request_count += 1
                    return html

            except Exception as e:
                if attempt < max_retries - 1:
                    retry_delay = 10 + (attempt * 10)
                    self.logger.warning(f"⚠️ 网络错误，{retry_delay}秒后重试: {e}")
                    time.sleep(retry_delay)
                else:
                    self.logger.error(f"❌ 获取失败: {e}")
                    return None

        return None

    def parse_list_page(self, html: str, batch: int, vehicle_type_name: str) -> List[Dict]:
        """解析列表页 - 提取基本信息"""
        vehicles = []

        # 提取车辆块
        vehicle_pattern = r'<li[^>]*>.*?<span class="car_name"><a href="([^"]+)"[^>]*>([^<]+)</a></span>.*?</li>'
        vehicle_matches = re.findall(vehicle_pattern, html, re.DOTALL)

        for idx, (detail_link, car_name) in enumerate(vehicle_matches, 1):
            vehicle = {
                '批次': batch,
                '燃料类型': vehicle_type_name,
                '车辆名称': car_name.strip(),
                '详情链接': f"{self.base_url}{detail_link}",
            }

            # 提取公告编号
            bh_match = re.search(r'bh=([^&"\']+)', detail_link)
            if bh_match:
                vehicle['公告编号'] = bh_match.group(1)

            # 从列表页提取更多信息
            detail_link_escaped = re.escape(detail_link)
            vehicle_block_pattern = rf'<li[^>]*>.*?<a href="{detail_link_escaped}".*?</li>'
            vehicle_block_match = re.search(vehicle_block_pattern, html, re.DOTALL)

            if vehicle_block_match:
                block = vehicle_block_match.group(0)

                # 提取公告型号
                model_pattern = r'<span class="car_model">([^<]+?)</span>'
                model_match = re.search(model_pattern, block)
                if model_match:
                    vehicle['公告型号'] = model_match.group(1).strip()

                # 提取生产企业
                company_patterns = [
                    r'车辆生产企业名称[：:]([^<\n]+?)(?:<|$)',
                    r'<label>生产企业[：:]([^<]+?)</label>',
                    r'<label>qymc[：:]([^<]+?)</label>',
                ]

                for pattern in company_patterns:
                    company_match = re.search(pattern, block)
                    if company_match:
                        vehicle['生产企业'] = company_match.group(1).strip()
                        break

            vehicles.append(vehicle)

        return vehicles

    def parse_detail_page_complete(self, html: str) -> Dict:
        """解析详情页 - 完整的96个字段"""
        detail_data = {}

        # 1. 解析生产企业信息表格（第一个表格）
        detail_data.update(self._parse_company_info(html))

        # 2. 解析车辆燃料参数
        detail_data.update(self._parse_fuel_params(html))

        # 3. 解析主要技术参数（左右两列表格）
        detail_data.update(self._parse_vehicle_params(html))

        # 4. 解析三电参数
        detail_data.update(self._parse_three_electric_params(html))

        # 5. 解析"其他"字段（原始文本）
        detail_data.update(self._parse_other_section(html))

        return detail_data

    def _clean_html_value(self, html: str) -> str:
        """
        清理HTML值，提取纯文本

        Args:
            html: 原始HTML内容

        Returns:
            清理后的纯文本
        """
        if not html:
            return ''

        # 移除<script>标签及其内容（JavaScript动态生成的内容无法提取）
        html = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)

        # 移除所有HTML标签
        html = re.sub(r'<[^>]+>', '', html)

        # 清理HTML实体（按顺序，避免重复替换）
        html = html.replace('&nbsp;', ' ')
        html = html.replace('&lt;', '<')
        html = html.replace('&gt;', '>')
        html = html.replace('&amp;', '&')
        html = html.replace('&quot;', '"')
        html = html.replace('&apos;', "'")
        html = html.replace('&times;', '×')
        html = html.replace('&#215;', '×')
        html = html.replace('&#039;', "'")
        html = html.replace('&#x27;', "'")
        html = html.replace('&#x3D;', '=')
        html = html.replace('&#061;', '=')

        # 清理多余的空格和换行符
        html = re.sub(r'\s+', ' ', html)
        html = html.strip()

        return html

    def _parse_company_info(self, html: str) -> Dict:
        """解析生产企业信息"""
        info = {}
        prefix = "生产企业信息_"

        # 找到生产企业信息的开始和结束位置
        section_start = html.find('生产企业信息')

        if section_start == -1:
            return info

        # 找到section结束（下一个section开始）
        # ✅ 修复：查找最近的下一个section
        # 注意：生产企业信息后面可能有"免检说明"、"公告状态"等section，要停止解析
        next_sections = [
            '免检说明',
            '公告状态',
            '主要技术参数',
            '车辆参数',
            '车辆燃料参数'
        ]
        section_end = len(html)
        for next_section in next_sections:
            pos = html.find(next_section, section_start + 1)
            if pos != -1 and pos < section_end:
                section_end = pos

        # 或者到表格结束
        table_end = html.find('</table>', section_start)
        if table_end != -1 and table_end < section_end:
            section_end = table_end

        section_html = html[section_start:section_end]

        # ✅ 修复：正确配对<th>和<td>（左右布局）
        tr_pattern = r'<tr[^>]*>.*?</tr>'
        for tr_match in re.finditer(tr_pattern, section_html, re.DOTALL):
            row_html = tr_match.group(0)

            # 提取所有<th>
            th_pattern = r'<th[^>]*>(.*?)</th>'
            ths = re.findall(th_pattern, row_html, re.DOTALL)

            # 提取所有<td>
            td_pattern = r'<td[^>]*>(.*?)</td>'
            tds = re.findall(td_pattern, row_html, re.DOTALL)

            # ✅ 跳过section标题行（在循环前检查）
            if ths and not tds:  # 只有<th>没有<td>，可能是标题行
                first_th_text = self._clean_html_value(ths[0]).strip()
                if first_th_text in ['生产企业信息', '企业信息', '免检说明', '公告状态']:
                    continue

            # 配对：th[i]对应td[i]
            for i in range(min(len(ths), len(tds))):
                key = self._clean_html_value(ths[i]).strip().rstrip('：:')

                # ✅ 去掉括号中的单位 (Kg)、(Km/h)、(mm)等 - 支持中英文括号
                key = re.sub(r'[\(（][^)\）]*[\)）]', '', key).strip()

                # ✅ 去除字段名后面的单位标识（KM、mm、Kg、Kw、Ah、V、Wh等）
                # ⚠️ 注意：不要包含"人"、"个"等，因为它们可能是字段名的一部分（如"联系人"）
                key = re.sub(r'(KM|km|MM|mm|KG|kg|KW|kw|AH|ah|V|v|WH|wh|％|%|／|/)$', '', key).strip()

                # 再次跳过section标题（保险）
                if key in ['生产企业信息', '企业信息', '免检说明', '公告状态']:
                    continue

                value = self._clean_html_value(tds[i])

                # 保存（即使为空也保存）
                if key:
                    info[prefix + key] = value

        return info

    def _parse_fuel_params(self, html: str) -> Dict:
        """解析车辆燃料参数"""
        info = {}
        prefix = "车辆燃料参数_"
        section_start = html.find('车辆燃料参数')

        if section_start == -1:
            return info

        # ✅ 修复：找到该section的结束位置（查找下一个section）
        # 在"车辆燃料参数"后面的section可能是：电芯参数、车辆制动参数、车辆底盘参数等
        next_sections = [
            '电芯参数',
            '车辆制动参数',
            '车辆底盘参数',
            '发动机参数',
            '反光标识参数',
            '其他',
            '更多图片'
        ]

        section_end = len(html)
        for next_section in next_sections:
            pos = html.find(next_section, section_start + 1)
            if pos != -1 and pos < section_end:
                section_end = pos

        section_html = html[section_start:section_end]

        # ✅ 修复：正确配对<th>和<td>（左右布局）
        tr_pattern = r'<tr[^>]*>.*?</tr>'
        for tr_match in re.finditer(tr_pattern, section_html, re.DOTALL):
            row_html = tr_match.group(0)

            # 提取所有<th>
            th_pattern = r'<th[^>]*>(.*?)</th>'
            ths = re.findall(th_pattern, row_html, re.DOTALL)

            # 提取所有<td>
            td_pattern = r'<td[^>]*>(.*?)</td>'
            tds = re.findall(td_pattern, row_html, re.DOTALL)

            # ✅ 跳过section标题行（在循环前检查）
            if ths and not tds:  # 只有<th>没有<td>，可能是标题行
                first_th_text = self._clean_html_value(ths[0]).strip()
                if first_th_text in ['车辆燃料参数', '燃料参数']:
                    continue

            # 配对：th[i]对应td[i]
            for i in range(min(len(ths), len(tds))):
                key = self._clean_html_value(ths[i]).strip().rstrip('：:')

                # ✅ 去掉括号中的单位 - 支持中英文括号
                key = re.sub(r'[\(（][^)\）]*[\)）]', '', key).strip()

                # ✅ 去除字段名后面的单位标识（KM、mm、Kg、Kw、Ah、V、Wh等）
                # ⚠️ 注意：不要包含"人"、"个"等，因为它们可能是字段名的一部分（如"联系人"）
                key = re.sub(r'(KM|km|MM|mm|KG|kg|KW|kw|AH|ah|V|v|WH|wh|％|%|／|/)$', '', key).strip()

                # 再次跳过section标题（保险）
                if key in ['车辆燃料参数', '燃料参数']:
                    continue

                value = self._clean_html_value(tds[i])

                # 保存（即使为空也保存）
                if key:
                    info[prefix + key] = value

        return info

    def _parse_vehicle_params(self, html: str) -> Dict:
        """解析主要技术参数（左右两列的表格）"""
        info = {}
        prefix = "主要技术参数_"

        # ✅ 修复：同时查找新旧两种section标题
        section_start = html.find('主要技术参数')
        if section_start == -1:
            section_start = html.find('车辆参数')  # 兼容旧网页

        if section_start == -1:
            return info

        # ✅ 修复：找到section结束（下一个section开始，而不是表格结束）
        # 在"主要技术参数"后面的section可能是：车辆燃料参数、车辆制动参数、车辆底盘参数等
        next_sections = [
            '车辆燃料参数',
            '车辆制动参数',
            '车辆底盘参数',
            '发动机参数',
            '反光标识参数',
            '其他',
            '更多图片'
        ]

        section_end = len(html)
        for next_section in next_sections:
            pos = html.find(next_section, section_start + 1)
            if pos != -1 and pos < section_end:
                section_end = pos

        # 或者到表格结束
        table_end = html.find('</table>', section_start)
        if table_end != -1 and table_end < section_end:
            section_end = table_end

        section_html = html[section_start:section_end]

        # ✅ 修复：正确配对<th>和<td>（左右布局）
        tr_pattern = r'<tr[^>]*>.*?</tr>'
        for tr_match in re.finditer(tr_pattern, section_html, re.DOTALL):
            row_html = tr_match.group(0)

            # 提取所有<th>
            th_pattern = r'<th[^>]*>(.*?)</th>'
            ths = re.findall(th_pattern, row_html, re.DOTALL)

            # 提取所有<td>
            td_pattern = r'<td[^>]*>(.*?)</td>'
            tds = re.findall(td_pattern, row_html, re.DOTALL)

            # ✅ 跳过section标题行（在循环前检查）
            if ths and not tds:  # 只有<th>没有<td>，可能是标题行
                first_th_text = self._clean_html_value(ths[0]).strip()
                if first_th_text in ['主要技术参数', '车辆参数', '法律法规']:
                    continue

            # 配对：th[i]对应td[i]
            for i in range(min(len(ths), len(tds))):
                key = self._clean_html_value(ths[i]).strip().rstrip('：:')

                # ✅ 去掉括号中的单位 - 支持中英文括号
                key = re.sub(r'[\(（][^)\）]*[\)）]', '', key).strip()

                # ✅ 去除字段名后面的单位标识（KM、mm、Kg、Kw、Ah、V、Wh等）
                # ⚠️ 注意：不要包含"人"、"个"等，因为它们可能是字段名的一部分（如"联系人"）
                key = re.sub(r'(KM|km|MM|mm|KG|kg|KW|kw|AH|ah|V|v|WH|wh|％|%|／|/)$', '', key).strip()

                # 再次跳过section标题和非参数字段（保险）
                if key in ['车辆参数', '法律法规'] or not key:
                    continue

                value = self._clean_html_value(tds[i])

                # 保存（即使为空也保存）
                if key:
                    info[prefix + key] = value

        return info

    def _parse_three_electric_params(self, html: str) -> Dict:
        """解析三电参数（电芯、电池、电机、电控）"""
        three_electric = {}

        # 定义section及其可能的关键词
        sections = [
            ('电芯参数', '电芯参数_'),  # ✅ 修改：输出为"电芯参数_"
            ('电池信息', '电池信息_'),
            ('电机信息', '电机信息_'),
            ('电控系统', '电控系统_')
        ]

        for section_name_cn, prefix in sections:
            # 尝试找到section标题（可能在<th>或普通文本中）
            section_patterns = [
                f'<th[^>]*>{section_name_cn}</th>',
                f'<th[^>]*>{section_name_cn}（',
                f'<td[^>]*>{section_name_cn}</td>',
                f'<td[^>]*>{section_name_cn}（',
                f'<b[^>]*>{section_name_cn}</b>',
                f'<strong[^>]*>{section_name_cn}</strong>',
                section_name_cn  # 纯文本匹配
            ]

            section_start = -1
            for pattern in section_patterns:
                pos = html.find(pattern.replace('<th[^>*</th>]', '<th>'))
                if pos != -1:
                    section_start = pos
                    break

            if section_start == -1:
                continue

            # 找到section结束（下一个section开始或表格结束）
            section_end = len(html)

            # ✅ 修复：查找下一个section（包括可能的其他section）
            all_next_sections = [
                # 三电参数的其他section
                '电池信息', '电机信息', '电控系统',
                # 可能出现在后面的其他section
                '更多图片',
                '交通部车辆燃料消耗量达标公告',
                '公告变更原因',
                '反光标识参数'
            ]

            # 当前section之后的section
            for next_section in all_next_sections:
                if next_section != section_name_cn:
                    # 尝试多种模式
                    for pattern in [
                        f'<th[^>]*>{next_section}</th>',
                        f'<th[^>]*>{next_section}（',
                        f'<td[^>]*>{next_section}</td>',
                        f'<td[^>]*>{next_section}（',
                        next_section
                    ]:
                        pos = html.find(pattern.replace('<th[^>*</th>]', '<th>'), section_start + 1)
                        if pos != -1 and pos < section_end:
                            section_end = pos
                            break

            # 或者到表格结束
            table_end = html.find('</table>', section_start)
            if table_end != -1 and table_end < section_end:
                section_end = table_end

            section_html = html[section_start:section_end]

            # ✅ 修复：正确配对<th>和<td>（左右布局）
            tr_pattern = r'<tr[^>]*>.*?</tr>'
            for tr_match in re.finditer(tr_pattern, section_html, re.DOTALL):
                row_html = tr_match.group(0)

                # 提取所有<th>
                th_pattern = r'<th[^>]*>(.*?)</th>'
                ths = re.findall(th_pattern, row_html, re.DOTALL)

                # 提取所有<td>
                td_pattern = r'<td[^>]*>(.*?)</td>'
                tds = re.findall(td_pattern, row_html, re.DOTALL)

                # ✅ 跳过section标题行（在循环前检查）
                skip_keywords = ['电芯参数', '电池信息', '电机信息', '电控系统',
                               '非公告数据', '仅供参考', '下面的三电参']
                if ths and not tds:  # 只有<th>没有<td>，可能是标题行
                    first_th_text = self._clean_html_value(ths[0]).strip()
                    if any(keyword in first_th_text for keyword in skip_keywords):
                        continue

                # 配对：th[i]对应td[i]
                for i in range(min(len(ths), len(tds))):
                    key = self._clean_html_value(ths[i]).strip().rstrip('：:')

                    # ✅ 去掉括号中的单位 (Kg)、(Km/h)、(mm)等
                    key = re.sub(r'\([^)]+\)', '', key).strip()

                    # ✅ 去除字段名后面的单位标识（KM、mm、Kg、Kw、Ah、V、Wh等）
                    # ⚠️ 注意：不要包含"人"、"个"等，因为它们可能是字段名的一部分（如"联系人"）
                    key = re.sub(r'(KM|km|MM|mm|KG|kg|KW|kw|AH|ah|V|v|WH|wh|％|%|／|/)$', '', key).strip()

                    # 再次跳过section标题本身和无关字段（保险）
                    if any(keyword in key for keyword in skip_keywords):
                        continue

                    value = self._clean_html_value(tds[i])

                    # 保存（即使为空也保存）
                    if key:
                        # 使用完整前缀区分不同section
                        full_key = f"{prefix}{key}"
                        three_electric[full_key] = value

        return three_electric

    def _parse_other_section(self, html: str) -> Dict:
        """解析"其他"字段 - 完整原始文本"""
        other_data = {}

        # 查找"其他"section
        section_start = html.find('其他</span>')
        if section_start == -1:
            # 尝试其他匹配模式
            section_start = html.find('>其他<')
            if section_start == -1:
                return other_data

        # 找到section结束（表格结束或下一个section）
        table_end = html.find('</table>', section_start)
        if table_end == -1:
            return other_data

        section_end = table_end

        # 提取"其他"section的内容
        section_html = html[section_start:section_end]

        # 匹配<p class="newline">标签中的内容
        content_pattern = r'<p[^>]*class="newline"[^>]*>(.*?)</p>'
        content_match = re.search(content_pattern, section_html, re.DOTALL | re.IGNORECASE)

        if content_match:
            # 提取原始文本
            raw_text = content_match.group(1)

            # ✅ 完整清理HTML标签
            raw_text = re.sub(r'<[^>]+>', '', raw_text)

            # ✅ 完整清理HTML实体（按顺序，避免重复替换）
            raw_text = raw_text.replace('&nbsp;', ' ')
            raw_text = raw_text.replace('&lt;', '<')
            raw_text = raw_text.replace('&gt;', '>')
            raw_text = raw_text.replace('&amp;', '&')
            raw_text = raw_text.replace('&quot;', '"')
            raw_text = raw_text.replace('&apos;', "'")
            raw_text = raw_text.replace('&times;', '×')
            raw_text = raw_text.replace('&#215;', '×')
            raw_text = raw_text.replace('&#039;', "'")
            raw_text = raw_text.replace('&#039;', "'")
            raw_text = raw_text.replace('&#x27;', "'")
            raw_text = raw_text.replace('&#x3D;', '=')
            raw_text = raw_text.replace('&#061;', '=')

            # 清理多余的空格和换行符
            raw_text = re.sub(r'\s+', ' ', raw_text)  # 多个空格合并为一个
            raw_text = raw_text.strip()

            # 保存原始文本
            if raw_text:
                other_data['其他'] = raw_text

        return other_data

    def download_vehicle_images(self, html: str, detail_url: str, vehicle_info: Dict) -> List[str]:
        """下载车辆图片"""
        images = []

        try:
            img_pattern = r'<img[^>]+src="([^"]+)"'
            img_matches = re.findall(img_pattern, html)

            bh = vehicle_info.get('公告编号', 'unknown')
            vehicle_name = vehicle_info.get('车辆名称', 'unknown')
            vehicle_name = re.sub(r'[<>:"/\\|?*]', '_', vehicle_name)[:30]

            downloaded = set()

            for idx, img_url in enumerate(img_matches, 1):
                # 过滤车辆图片
                if not any(x in img_url.lower() for x in ['ufileos', 'vehicle', 'car', 'photo']):
                    continue

                # 构造完整URL
                if img_url.startswith('//'):
                    img_url = 'https:' + img_url
                elif img_url.startswith('/'):
                    img_url = self.base_url + img_url

                img_url_clean = img_url.split('?')[0]

                if img_url_clean in downloaded:
                    continue
                downloaded.add(img_url_clean)

                try:
                    ext = os.path.splitext(img_url_clean)[1] or '.jpg'
                    filename = f"{bh}_{vehicle_name}_{idx}{ext}"
                    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
                    filepath = os.path.join(self.image_dir, filename)

                    req = urllib.request.Request(img_url, headers={
                        'User-Agent': 'Mozilla/5.0',
                        'Referer': detail_url
                    })

                    with urllib.request.urlopen(req, timeout=15) as response:
                        with open(filepath, 'wb') as f:
                            f.write(response.read())

                    images.append(filepath)
                    self.logger.debug(f"下载图片: {filename}")

                except Exception as e:
                    self.logger.warning(f"图片下载失败 {img_url}: {e}")
                    continue

        except Exception as e:
            self.logger.warning(f"图片处理失败: {e}")

        return images

    def crawl_single_batch_list(self, batch: int, vehicle_type_code: str, vehicle_type_name: str) -> List[Dict]:
        """爬取单个批次的列表页"""
        url = f"{self.base_url}/qcggs"
        params = {
            'ggxh': '',
            'ggpc': str(batch),
            'zwpp': '',
            'clmc': vehicle_type_code,  # 车辆类型代码（34种之一）
            'fdjxh': '',
            'qymc': '',
            'cph': '',
            'rylx': '',  # 燃料类型留空
            'viewtype': '0'
        }
        list_url = f"{url}?{urllib.parse.urlencode(params)}"

        self.logger.info(f"爬取列表页: 批次{batch} [{vehicle_type_name}]")

        html = self.fetch_with_session(list_url)

        if not html:
            self.logger.warning(f"批次{batch} [{vehicle_type_name}] - 获取列表页失败")
            return []

        vehicles = self.parse_list_page(html, batch, vehicle_type_name)
        self.logger.info(f"批次{batch} [{vehicle_type_name}] - 找到 {len(vehicles)} 辆")

        return vehicles

    def crawl_vehicle_detail(self, vehicle: Dict) -> Optional[Dict]:
        """爬取单个车辆的详情页"""
        detail_url = vehicle.get('详情链接')

        if not detail_url:
            return None

        try:
            # 使用列表页URL作为referer
            list_url = f"{self.base_url}/qcggs"
            html = self.fetch_with_session(detail_url, referer=list_url)

            if not html:
                self.logger.warning(f"详情页获取失败: {vehicle.get('公告编号', 'unknown')}")
                return None

            # 解析详情页
            detail_data = self.parse_detail_page_complete(html)

            # 下载图片
            images = self.download_vehicle_images(html, detail_url, vehicle)
            if images:
                detail_data['车辆图片'] = images

            # 合并列表页数据和详情页数据
            complete_data = {**vehicle, **detail_data}
            complete_data['详情页URL'] = detail_url
            complete_data['爬取时间'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            self.logger.debug(f"详情页解析完成: {vehicle.get('公告编号', 'unknown')}, 字段数: {len(detail_data)}")

            return complete_data

        except Exception as e:
            self.logger.error(f"详情页爬取失败 {vehicle.get('公告编号', 'unknown')}: {e}")
            return None

    def run_full_crawl(self, start_batch: int = 1, end_batch: int = 403, crawl_detail: bool = True, enable_working_hours: bool = False, auto_export_excel: bool = True):
        """执行全量爬取"""
        self.logger.info("=" * 70)
        self.logger.info(f"开始全量爬取: 批次 {start_batch}-{end_batch}")
        self.logger.info(f"详情页爬取: {'启用' if crawl_detail else '禁用'}")
        self.logger.info("=" * 70)

        print("=" * 70)
        print("🛡️ 增强版生产环境爬虫 - 安全模式")
        print("=" * 70)
        print(f"批次范围: {start_batch} - {end_batch}")
        print(f"爬取详情: {'是 (完整96字段)' if crawl_detail else '否 (仅列表页6字段)'}")
        print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"日志文件: {self.log_file}")
        print()
        print("⚙️ 安全配置:")
        print(f"  - 详情页延迟: {self.detail_delay_min}-{self.detail_delay_max} 秒")
        print(f"  - 批次间延迟: {self.batch_delay_min}-{self.batch_delay_max} 秒")
        print(f"  - 休息策略: 每 {self.rest_interval} 批次休息 {self.rest_duration} 秒")
        print(f"  - 工作时间限制: {'启用 (仅9:00-18:00)' if enable_working_hours else '禁用 (24小时运行)'}")
        print(f"  - 自动导出Excel: {'是' if auto_export_excel else '否 (可后续手动导出)'}")
        print("=" * 70)
        print()

        # 🔥 预热：先访问首页建立会话
        self.warm_up()

        # 加载状态
        state = self.load_state()
        all_vehicles = self.load_data()

        self.logger.info(f"已有数据: {len(all_vehicles)} 辆")
        self.logger.info(f"已完成批次: {len(state['completed_batches'])}")

        print(f"📊 已有数据: {len(all_vehicles)} 辆")
        print(f"✓ 已完成批次: {len(state['completed_batches'])}")

        # 车辆类型（34种新能源汽车）
        vehicle_types = [
            # 纯电动系列（15种）
            # {'code': '1', 'name': '纯电动汽车'},  # ❌ 已爬取，暂时注释
            {'code': '294', 'name': '纯电动客车'},
            {'code': '295', 'name': '纯电动救护车'},
            {'code': '296', 'name': '纯电动载货车'},
            {'code': '297', 'name': '纯电动洒水车'},
            {'code': '300', 'name': '纯电动教练车'},
            {'code': '284', 'name': '换电式纯电动轿车'},
            {'code': '285', 'name': '换电式纯电动自卸汽车'},
            {'code': '286', 'name': '换电式纯电动厢式运输车'},
            {'code': '287', 'name': '换电式纯电动多用途乘用车'},
            {'code': '288', 'name': '换电式纯电动自卸式垃圾车'},
            {'code': '289', 'name': '换电式纯电动半挂牵引车'},
            {'code': '290', 'name': '换电式纯电动混凝土搅拌运输车'},
            {'code': '291', 'name': '换电式纯电动福祉多用途乘用车'},
            {'code': '182', 'name': '两用燃料汽车'},
            # 混合动力系列（4种）
            {'code': '2', 'name': '混合动力电动汽车'},
            {'code': '3', 'name': '插电式混合动力汽车'},
            {'code': '4', 'name': '增程式混合动力汽车'},
            {'code': '5', 'name': '燃料式电池汽车'},
            # 插电式混合动力专用车型（14种）
            {'code': '301', 'name': '插电式混合动力冷藏车'},
            {'code': '302', 'name': '插电式混合动力宣传车'},
            {'code': '303', 'name': '插电式混合动力清障车'},
            {'code': '304', 'name': '插电式混合动力扫路车'},
            {'code': '305', 'name': '插电式混合动力检测车'},
            {'code': '306', 'name': '插电式混合动力救护车'},
            {'code': '307', 'name': '插电式混合动力运钞车'},
            {'code': '308', 'name': '插电式混合动力房车'},
            {'code': '309', 'name': '插电式混合动力城市客车'},
            {'code': '317', 'name': '插电式混合动力垃圾车'},
            {'code': '318', 'name': '插电式混合动力牵引车'},
            {'code': '319', 'name': '插电式混合动力载货车'},
            {'code': '320', 'name': '插电式混合动力汽车起重机'},
            {'code': '321', 'name': '插电式混合动力厢式运输车'},
            {'code': '322', 'name': '插电式混合动力混凝土泵车'},
            {'code': '323', 'name': '插电式混合动力绿化喷洒车'},
            {'code': '324', 'name': '插电式混合动力多用途乘用车'},
            {'code': '325', 'name': '甲醇插电式增程混合动力车'},
            {'code': '326', 'name': '插电式混合动力仓栅式运输车'},
            {'code': '327', 'name': '插电式混合动力混凝土搅拌运输车'},
            {'code': '329', 'name': '插电式混合动力自卸汽车'},
        ]

        total_tasks = (end_batch - start_batch + 1) * len(vehicle_types)
        completed_tasks = len(state['completed_batches'])

        # 🔥 随机化批次顺序（避免顺序访问被检测）
        batch_list = list(range(start_batch, end_batch + 1))
        random.shuffle(batch_list)
        self.logger.info(f"🎲 已随机化批次顺序，避免顺序访问被检测")
        self.logger.info(f"批次顺序前10个: {batch_list[:10]}... (共{len(batch_list)}个批次)")
        print(f"🎲 已随机化批次顺序，避免顺序访问被检测")
        print(f"批次顺序前10个: {batch_list[:10]}... (共{len(batch_list)}个批次)")
        print()

        # 开始爬取
        for batch in batch_list:
            # 🔥 检查工作时间（如果启用）
            if enable_working_hours and not self.is_working_hours():
                print("\n" + "=" * 70)
                print("⏰ 非工作时间（仅9:00-18:00工作）")
                print("💡 程序将安全退出")
                print("💡 你可以在工作时间重新运行程序，它会自动跳过已完成的批次")
                print(f"📊 当前进度: {completed_tasks}/{total_tasks} ({(completed_tasks/total_tasks)*100:.1f}%)")
                print(f"📦 已采集车辆: {len(all_vehicles)} 辆")
                print("=" * 70)
                self.save_data(all_vehicles)
                state['total_vehicles'] = len(all_vehicles)
                self.save_state(state)
                return all_vehicles

            for vehicle_type in vehicle_types:
                task_key = f"{batch}-{vehicle_type['code']}"

                # 跳过已完成
                if task_key in state['completed_batches']:
                    continue

                try:
                    # 1. 爬取列表页（燃料类型留空）
                    vehicles = self.crawl_single_batch_list(batch, vehicle_type['code'], vehicle_type['name'])

                    if not vehicles:
                        # ✅ 详细失败记录
                        fail_record = {
                            'task_key': task_key,
                            'batch': batch,
                            'vehicle_type': vehicle_type['code'],
                            'vehicle_type_name': vehicle_type['name'],
                            'fail_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'fail_reason': '列表页无数据',
                            'error_type': 'NoData'
                        }
                        state['failed_batches'].append(fail_record)
                        continue

                    # 2. 爬取详情页（可选）
                    if crawl_detail:
                        batch_vehicles_complete = []
                        for idx, vehicle in enumerate(vehicles, 1):
                            print(f"    [{idx}/{len(vehicles)}] {vehicle.get('车辆名称', 'unknown')} - {vehicle.get('公告编号', 'unknown')}")

                            detail_data = self.crawl_vehicle_detail(vehicle)
                            if detail_data:
                                batch_vehicles_complete.append(detail_data)

                                # ✅ 每5辆车保存一次（防止中断时数据丢失）
                                if len(batch_vehicles_complete) % 5 == 0:
                                    # 保存本次批次已爬取的车辆
                                    temp_batch = batch_vehicles_complete[-5:]
                                    all_vehicles.extend(temp_batch)
                                    self.save_data(all_vehicles)
                                    state['total_vehicles'] = len(all_vehicles)
                                    self.save_state(state)
                                    print(f"    ✓ 进度已保存 ({len(all_vehicles)}辆总计)")

                            # 详情页之间延迟（更长，避免被封）
                            if idx < len(vehicles):
                                delay = random.uniform(self.detail_delay_min, self.detail_delay_max)
                                print(f"    ⏱️  等待 {delay:.1f} 秒...")
                                time.sleep(delay)

                        # 批次完成后的最终保存
                        if batch_vehicles_complete:
                            # 保存剩余的车辆（如果有的话）
                            remaining = len(batch_vehicles_complete) % 5
                            if remaining != 0:
                                all_vehicles.extend(batch_vehicles_complete[-remaining:])

                            # 更新状态
                            state['completed_batches'].append(task_key)
                            completed_tasks += 1

                            # 保存
                            self.save_data(all_vehicles)
                            state['total_vehicles'] = len(all_vehicles)
                            self.save_state(state)

                            progress = (completed_tasks / total_tasks) * 100
                            print(f"    进度: {completed_tasks}/{total_tasks} ({progress:.1f}%) | 累计: {len(all_vehicles)} 辆")
                        else:
                            # ✅ 详细失败记录
                            fail_record = {
                                'task_key': task_key,
                                'batch': batch,
                                'vehicle_type': vehicle_type['code'],
                                'vehicle_type_name': vehicle_type['name'],
                                'fail_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'fail_reason': '详情页无数据',
                                'error_type': 'NoDetailData'
                            }
                            state['failed_batches'].append(fail_record)
                    else:
                        # 不爬详情页，直接保存列表页数据
                        all_vehicles.extend(vehicles)
                        state['completed_batches'].append(task_key)
                        completed_tasks += 1

                        self.save_data(all_vehicles)
                        state['total_vehicles'] = len(all_vehicles)
                        self.save_state(state)

                        progress = (completed_tasks / total_tasks) * 100
                        print(f"    进度: {completed_tasks}/{total_tasks} ({progress:.1f}%) | 累计: {len(all_vehicles)} 辆")

                    # 批次间延迟（更长，更随机）
                    base_delay = random.uniform(self.batch_delay_min, self.batch_delay_max)
                    print(f"  ⏱️ 批次间延迟 {base_delay:.1f} 秒...")
                    time.sleep(base_delay)

                    # 定期休息（更频繁，更长时间）
                    if completed_tasks > 0 and completed_tasks % self.rest_interval == 0:
                        self.logger.info(f"已爬取{completed_tasks}个任务，休息{self.rest_duration}秒...")
                        print(f"\n{'='*70}")
                        print(f"💦 已爬取 {completed_tasks} 个任务")
                        print(f"⏸️  休息 {self.rest_duration} 秒（避免触发反爬）")
                        print(f"{'='*70}\n")
                        time.sleep(self.rest_duration)
                        print(f"✓ 休息完成，继续爬取\n")

                except KeyboardInterrupt:
                    print("\n\n" + "="*70)
                    print("⚠️  用户中断！保存当前进度...")
                    print("="*70)

                    # ✅ 修复：保存当前批次已爬取的数据
                    saved_count = 0
                    if 'batch_vehicles_complete' in locals() and batch_vehicles_complete:
                        # 找出还没有保存到all_vehicles的车辆
                        # 计算batch_vehicles_complete中哪些车辆还没有保存
                        saved_in_batch = len(batch_vehicles_complete) // 5 * 5  # 已保存的数量
                        remaining = batch_vehicles_complete[saved_in_batch:]  # 剩余未保存的

                        if remaining:
                            all_vehicles.extend(remaining)
                            saved_count = len(remaining)

                    # 保存所有数据
                    self.save_data(all_vehicles)
                    state['total_vehicles'] = len(all_vehicles)
                    self.save_state(state)

                    print(f"✓ 已保存 {len(all_vehicles)} 辆车的数据")
                    if saved_count > 0:
                        print(f"✓ 本次中断保存了 {saved_count} 辆未保存的数据")
                    print("="*70)

                    # ✅ 询问是否导出Excel
                    print("\n💾  数据已保存到文件")
                    print("💡 您可以：")
                    print("  • 选择菜单 '5. 导出Excel' 导出当前数据")
                    print("  • 选择菜单 '4. 继续任务' 继续未完成的爬取")
                    print()

                    return

                except Exception as e:
                    print(f"    ✗ 错误: {e}")
                    self.logger.error(f"批次{batch}错误: {e}")

                    # ✅ 详细失败记录
                    fail_record = {
                        'task_key': task_key,
                        'batch': batch,
                        'vehicle_type': vehicle_type['code'],
                        'vehicle_type_name': vehicle_type['name'],
                        'fail_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'fail_reason': str(e)[:200],  # 限制长度
                        'error_type': type(e).__name__
                    }
                    state['failed_batches'].append(fail_record)
                    self.logger.error(f"失败详情: {fail_record}")

        # 完成
        self.logger.info("=" * 70)
        self.logger.info("爬取完成")
        self.logger.info(f"总计: {len(all_vehicles)} 辆车")
        self.logger.info(f"成功: {len(state['completed_batches'])} 个任务")
        self.logger.info(f"失败: {len(state['failed_batches'])} 个任务")
        self.logger.info("=" * 70)

        print("=" * 70)
        print("✅ 爬取完成！")
        print(f"✓ 总计: {len(all_vehicles)} 辆车")
        print(f"✓ 成功: {len(state['completed_batches'])} 个任务")
        print(f"✗ 失败: {len(state['failed_batches'])} 个任务")
        print(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 70)
        print()

        # 导出Excel（如果启用）
        if all_vehicles and auto_export_excel:
            self.export_to_excel(all_vehicles)
        elif all_vehicles and not auto_export_excel:
            print("💡 跳过Excel导出")
            print("💡 可使用选项5手动导出Excel")
            print()

        return all_vehicles

    def export_to_excel(self, vehicles: List[Dict]):
        """导出到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            import openpyxl.utils  # ✅ 新增：导入utils模块
        except ImportError:
            print("⚠️  需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
            return

        filename = os.path.basename(self.excel_file)
        print(f"📊 正在导出Excel: {filename}")

        # ✅ 使用预定义的有序字段列表（按Section顺序）
        # 同时收集实际存在的额外字段（如果有）
        all_fields = set()
        for vehicle in vehicles:
            all_fields.update(vehicle.keys())

        # 合并预定义的有序字段和实际字段
        ordered_fields = list(self.ORDERED_FIELDS)
        for field in ordered_fields:
            if field in all_fields:
                all_fields.remove(field)

        # 添加预定义列表中不存在的额外字段（如果有）
        extra_fields = sorted(all_fields)
        sorted_fields = ordered_fields + extra_fields

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "车辆完整数据"

        # 写入表头
        for col_idx, field_name in enumerate(sorted_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row_idx, vehicle in enumerate(vehicles, 2):
            for col_idx, field_name in enumerate(sorted_fields, 1):
                value = vehicle.get(field_name, '')

                # 处理列表类型（如图片路径）
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                elif value is None:
                    value = ''

                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # ✅ 特殊处理"其他"字段（长文本）
                if field_name == '其他' and value:
                    # 启用自动换行和文本换行
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',  # 长文本顶部对齐更好
                        wrap_text=True,  # ✅ 启用自动换行
                        shrink_to_fit=False  # 不自动缩小字体
                    )
                else:
                    # 普通字段保持原有对齐
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # 冻结首行
        ws.freeze_panes = 'A2'

        # ✅ 设置列宽（特殊处理"其他"字段）
        for col_idx, field_name in enumerate(sorted_fields, 1):
            # 获取列字母（如A, B, C, ... AA, AB, ...）
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            if field_name == '其他':
                # "其他"字段设置更宽的列宽（100个字符宽）
                ws.column_dimensions[col_letter].width = 100
            elif field_name in ['车辆图片', '详情页URL']:
                # URL和图片路径也设置较宽
                ws.column_dimensions[col_letter].width = 50
            # 其他列使用自动宽度（不设置，让Excel自动调整）

        # 保存
        wb.save(self.excel_file)
        file_size = os.path.getsize(self.excel_file) / 1024 / 1024
        print(f"✓ Excel已保存: {self.excel_file}")
        print(f"  文件大小: {file_size:.2f} MB")
        print(f"  记录数: {len(vehicles)}")
        print(f"  字段数: {len(sorted_fields)}")


def main():
    """主函数"""
    crawler = EnhancedProductionCrawler()

    print("\n" + "=" * 70)
    print("增强版全量数据爬取 - 优化版")
    print("=" * 70)
    print("\n📋 选项:")
    print("1. 【完整模式】爬取所有批次+详情页 (96个字段，非常慢)")
    print("2. 【快速模式】仅爬取列表页 (6个字段，快速)")
    print("3. 【测试模式】爬取单个批次测试 (385批次) ⚡快速延迟")
    print("4. 【继续任务】继续上次中断的任务")
    print("5. 【导出Excel】仅导出已有数据为Excel")
    print("6. 【分段爬取】每天爬50批次 (更安全，避免被封)")
    print("7. 【批次范围】指定批次范围 - 起始到结束 (如: 100-150)")
    print("8. 【指定批次】指定单个批次号 (如: 385)")
    print()

    choice = input("请选择 (1-8): ").strip()

    if choice == '1':
        print("\n你选择了：【完整模式】爬取所有批次+详情页 (96个字段)")

        # 选择延迟配置
        profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=False)
        profile = crawler.apply_delay_profile(profile_key)

        # 显示时间估算
        print("\n" + "="*70)
        print("📊 时间估算")
        print("="*70)
        print("爬取范围：1-403批次（全量）")
        print(f"爬取字段：96个（完整版）")
        print(f"延迟配置：{profile['emoji']} {profile['name']}（{profile['detail_min']}-{profile['detail_max']}秒 / {profile['batch_min']}-{profile['batch_max']}秒）")
        print()
        print("计算明细：")
        print("• 预计车辆数：约24,180-48,360辆")
        print(f"• 详情页总耗时：{crawler.estimate_time(403, 24180, profile)}")
        print(f"• 批次间总耗时：{crawler.estimate_time(403, 0, profile)}")
        print()
        estimated_total = crawler.estimate_time(403, 24180, profile)
        print(f"⏱️  总预计耗时：{estimated_total}")
        print("="*70)
        print(f"✅  这是\"{profile['risk']}\"等级")
        print()

        # 其他配置选项
        print("配置选项:")
        enable_work_hours = crawler.ask_working_hours()
        auto_export_excel = crawler.ask_export_excel()

        confirm = input("\n确认继续？(yes/no): ").strip().lower()
        if confirm == 'yes':
            crawler.run_full_crawl(start_batch=1, end_batch=403, crawl_detail=True, enable_working_hours=enable_work_hours, auto_export_excel=auto_export_excel)
        else:
            print("已取消")
    elif choice == '2':
        print("\n你选择了：【快速模式】仅爬取列表页 (6个字段，快速)")
        print("\n⚠️  注意：快速模式只爬取列表页，不爬取详情页")
        print("⚡  速度很快，不需要延迟配置")

        # 显示时间估算
        print("\n" + "="*70)
        print("📊 时间估算")
        print("="*70)
        print("爬取范围：1-403批次（全量）")
        print("爬取字段：6个（列表页字段）")
        print()
        print("计算明细：")
        print("• 批次总数：403个")
        print("• 列表页请求：806次（403纯电动 + 403混合动力）")
        print("• 每次请求约1秒")
        print()
        print("⏱️  总预计耗时：约13-15分钟")
        print("="*70)
        print()

        crawler.run_full_crawl(start_batch=1, end_batch=403, crawl_detail=False)
    elif choice == '3':
        print("\n你选择了：【测试模式】爬取单个批次测试 (385批次)")
        print("\n⚡  测试模式 - 推荐使用快速延迟验证修复效果")

        # ✅ 让用户选择延迟配置
        print("\n" + "⏱️ " * 35)
        print("测试模式 - 延迟配置")
        print("⏱️ " * 35)
        print("\n💡 测试模式推荐使用快速延迟（3-8秒），不频繁测试不会被检测\n")

        profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=False)
        profile = crawler.apply_delay_profile(profile_key)

        # 显示时间估算
        print("\n" + "="*70)
        print("📊 时间估算")
        print("="*70)
        print("爬取范围：385批次")
        print("爬取字段：96个（完整版）")
        print(f"延迟配置：{profile['emoji']} {profile['name']}（{profile['detail_min']}-{profile['detail_max']}秒）")
        print()
        print("计算明细：")
        print("• 预计车辆数：约60辆（仅纯电动）")
        # 估算时间（单批次约60辆车）
        estimated_minutes = int(60 * ((profile['detail_min'] + profile['detail_max']) / 2) / 60)
        print(f"• 详情页总耗时：约{estimated_minutes}分钟")
        print("• 无批次间延迟（单批次）")
        print()
        print(f"⏱️  总预计耗时：约{estimated_minutes}分钟")
        print("="*70)
        print(f"💡  这是\"{profile['risk']}\"等级")
        print("💡  提示：测试模式用于快速验证代码修复效果")
        print()

        crawler.run_full_crawl(start_batch=385, end_batch=385, crawl_detail=True)
    elif choice == '4':
        print("\n你选择了：【继续任务】继续上次中断的任务")

        # 加载状态
        state = crawler.load_state()
        all_vehicles = crawler.load_data()

        print(f"📊 已有数据：{len(all_vehicles)}辆")
        print(f"✓ 已完成批次：{len(state['completed_batches'])}个")

        if len(state['completed_batches']) > 0:
            print("\n剩余批次：401个（已完成2个）")

        # 选择延迟配置（不包括极速）
        profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=True)
        profile = crawler.apply_delay_profile(profile_key)

        # 显示时间估算
        remaining_batches = 403 - len(state['completed_batches'])
        print("\n" + "="*70)
        print("📊 时间估算")
        print("="*70)
        print(f"爬取范围：1-403批次")
        print(f"剩余批次：{remaining_batches}个")
        print("爬取字段：96个（完整版）")
        print(f"延迟配置：{profile['emoji']} {profile['name']}（{profile['detail_min']}-{profile['detail_max']}秒 / {profile['batch_min']}-{profile['batch_max']}秒）")
        print()
        estimated_remaining = crawler.estimate_time(remaining_batches, 24060, profile)
        print(f"⏱️  预计剩余耗时：{estimated_remaining}")
        print("="*70)
        print("💡  提示：程序会自动跳过已完成的批次")
        print()

        enable_work_hours = crawler.ask_working_hours()
        auto_export_excel = crawler.ask_export_excel()
        crawler.run_full_crawl(start_batch=1, end_batch=403, crawl_detail=True, enable_working_hours=enable_work_hours, auto_export_excel=auto_export_excel)
    elif choice == '5':
        all_vehicles = crawler.load_data()
        if all_vehicles:
            crawler.export_to_excel(all_vehicles)
        else:
            print("没有数据可导出")
    elif choice == '6':
        # 分段爬取模式
        print("\n" + "=" * 70)
        print("分段爬取模式 - 降低被封风险")
        print("=" * 70)
        print("\n📅 分段方案:")
        print("  第1段: 批次 1-50    (50个批次)")
        print("  第2段: 批次 51-100  (50个批次)")
        print("  第3段: 批次 101-150 (50个批次)")
        print("  第4段: 批次 151-200 (50个批次)")
        print("  第5段: 批次 201-250 (50个批次)")
        print("  第6段: 批次 251-300 (50个批次)")
        print("  第7段: 批次 301-350 (50个批次)")
        print("  第8段: 批次 351-403 (53个批次)")
        print("\n💡 建议:")
        print("  - 每天只爬1-2段")
        print("  - 给IP充足的休息时间（隔12-24小时）")
        print("  - 仅在工作时间爬取（9:00-18:00）")
        print("=" * 70)
        print()

        segment_map = {
            '1': (1, 50),
            '2': (51, 100),
            '3': (101, 150),
            '4': (151, 200),
            '5': (201, 250),
            '6': (251, 300),
            '7': (301, 350),
            '8': (351, 403),
        }

        segment = input("选择爬取第几段 (1-8): ").strip()

        if segment in segment_map:
            start, end = segment_map[segment]
            batch_count = end - start + 1

            # 选择延迟配置（不包括极速和超保守）
            profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=True)
            # 如果选的是超保守，重新选择
            if profile_key == '5':
                print("\n⚠️  分段爬取建议不要使用超保守模式（太慢）")
                profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=True)
            profile = crawler.apply_delay_profile(profile_key)

            # 显示时间估算
            print("\n" + "="*70)
            print("📊 时间估算")
            print("="*70)
            print(f"爬取范围：{start}-{end}批次（第{segment}段）")
            print(f"爬取字段：96个（完整版）")
            print(f"延迟配置：{profile['emoji']} {profile['name']}（{profile['detail_min']}-{profile['detail_max']}秒 / {profile['batch_min']}-{profile['batch_max']}秒）")
            print()
            print("计算明细：")
            print(f"• 预计车辆数：约{batch_count * 60}辆（{batch_count}批次 × 60辆）")
            print(f"• 详情页总耗时：{crawler.estimate_time(batch_count, batch_count * 60, profile)}")
            print(f"• 批次间总耗时：{crawler.estimate_time(batch_count, 0, profile)}")
            print()
            print(f"⏱️  总预计耗时：{crawler.estimate_time(batch_count, batch_count * 60, profile)}")
            print("="*70)
            print("💡  建议：每天爬1-2段，给IP充足休息时间")
            print()

            enable_work_hours = crawler.ask_working_hours()
            auto_export_excel = crawler.ask_export_excel()

            confirm = input("确认开始？(yes/no): ").strip().lower()
            if confirm == 'yes':
                crawler.run_full_crawl(start_batch=start, end_batch=end, crawl_detail=True, enable_working_hours=enable_work_hours, auto_export_excel=auto_export_excel)
            else:
                print("已取消")
        else:
            print("❌ 无效的选择")
    elif choice == '7':
        # 自定义批次范围
        print("\n你选择了：【批次范围】指定批次范围 - 起始到结束 (如: 100-150)")
        print("\n" + "=" * 70)
        print("批次范围爬取")
        print("=" * 70)
        print("\n💡 提示:")
        print("  - 批次范围: 1-403")
        print("  - 建议: 每次不超过50个批次")
        print("  - 可以指定任意批次范围")
        print("=" * 70)
        print()

        try:
            start_batch = input("请输入起始批次: ").strip()
            end_batch = input("请输入结束批次: ").strip()

            start_batch = int(start_batch)
            end_batch = int(end_batch)

            if not (1 <= start_batch <= 403) or not (1 <= end_batch <= 403):
                print("❌ 批次范围必须在 1-403 之间")
            elif start_batch > end_batch:
                print("❌ 起始批次不能大于结束批次")
            else:
                batch_count = end_batch - start_batch + 1
                print(f"\n批次范围：{start_batch}-{end_batch}")
                print(f"批次数量：{batch_count}个批次")

                # 选择延迟配置（不包括极速和超保守）
                profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=True)
                # 如果选的是超保守，重新选择
                if profile_key == '5':
                    print("\n⚠️  小范围爬取建议不要使用超保守模式（太慢）")
                    profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=True)
                profile = crawler.apply_delay_profile(profile_key)

                # 显示时间估算
                print("\n" + "="*70)
                print("📊 时间估算")
                print("="*70)
                print(f"爬取范围：{start_batch}-{end_batch}批次（{batch_count}个批次）")
                print("爬取字段：96个（完整版）")
                print(f"延迟配置：{profile['emoji']} {profile['name']}（{profile['detail_min']}-{profile['detail_max']}秒 / {profile['batch_min']}-{profile['batch_max']}秒）")
                print()
                print("计算明细：")
                print(f"• 预计车辆数：约{batch_count * 60}辆")
                print(f"• 详情页总耗时：{crawler.estimate_time(batch_count, batch_count * 60, profile)}")
                print(f"• 批次间总耗时：{crawler.estimate_time(batch_count, 0, profile)}")
                print()
                print(f"⏱️  总预计耗时：{crawler.estimate_time(batch_count, batch_count * 60, profile)}")
                print("="*70)
                print()

                enable_work_hours = crawler.ask_working_hours()
                auto_export_excel = crawler.ask_export_excel()

                # 询问是否爬取详情页
                crawl_detail = input("是否爬取详情页？(96个字段) [y/n]: ").strip().lower()
                enable_crawl_detail = crawl_detail == 'y'

                if enable_crawl_detail:
                    print(f"✓ 将爬取详情页（完整96字段）")
                else:
                    print(f"✓ 仅爬取列表页（6个字段，快速）")
                print()

                confirm = input("确认开始？(yes/no): ").strip().lower()
                if confirm == 'yes':
                    crawler.run_full_crawl(
                        start_batch=start_batch,
                        end_batch=end_batch,
                        crawl_detail=enable_crawl_detail,
                        enable_working_hours=enable_work_hours,
                        auto_export_excel=auto_export_excel
                    )
                else:
                    print("已取消")

        except ValueError:
            print("❌ 输入无效，请输入数字")
    elif choice == '8':
        # 单个批次爬取
        print("\n你选择了：【指定批次】指定单个批次号 (如: 385)")
        print("\n" + "=" * 70)
        print("指定单个批次爬取")
        print("=" * 70)
        print("\n💡 提示:")
        print("  - 批次号: 1-403")
        print("  - 适合爬取某个特定批次")
        print("=" * 70)
        print()

        try:
            batch_num = input("请输入批次号: ").strip()
            batch_num = int(batch_num)

            if not (1 <= batch_num <= 403):
                print("❌ 批次号必须在 1-403 之间")
            else:
                print(f"\n批次：{batch_num}")

                # 选择延迟配置
                profile_key = crawler.select_delay_profile(include_custom=True, exclude_fastest=False)
                profile = crawler.apply_delay_profile(profile_key)

                # 显示时间估算
                print("\n" + "="*70)
                print("📊 时间估算")
                print("="*70)
                print(f"爬取范围：{batch_num}批次（单批次）")
                print("爬取字段：96个（完整版）")
                print(f"延迟配置：{profile['emoji']} {profile['name']}（{profile['detail_min']}-{profile['detail_max']}秒）")
                print()
                print("计算明细：")
                print("• 纯电动：约60辆 × 平均延迟")
                print("• 混合动力：约0辆")
                print("• 无批次间延迟（单批次）")
                print("• 无休息时间")
                print()
                # 估算时间（单批次约60辆车）
                print(f"⏱️  总预计耗时：约{int(60 * ((profile['detail_min'] + profile['detail_max']) / 2) / 60)}分钟")
                print("="*70)
                print("💡  提示：单批次爬取，不会有批次间延迟")
                print()

                enable_work_hours = crawler.ask_working_hours()
                auto_export_excel = crawler.ask_export_excel()

                # 询问是否爬取详情页
                crawl_detail = input("是否爬取详情页？(96个字段) [y/n]: ").strip().lower()
                enable_crawl_detail = crawl_detail == 'y'

                if enable_crawl_detail:
                    print(f"✓ 将爬取详情页（完整96字段）")
                else:
                    print(f"✓ 仅爬取列表页（6个字段，快速）")
                print()

                confirm = input("确认开始？(yes/no): ").strip().lower()
                if confirm == 'yes':
                    crawler.run_full_crawl(
                        start_batch=batch_num,
                        end_batch=batch_num,
                        crawl_detail=enable_crawl_detail,
                        enable_working_hours=enable_work_hours,
                        auto_export_excel=auto_export_excel
                    )
                else:
                    print("已取消")

                confirm = input("确认开始? (yes/no): ").strip().lower()
                if confirm == 'yes':
                    crawler.run_full_crawl(
                        start_batch=batch_num,
                        end_batch=batch_num,
                        crawl_detail=enable_crawl_detail,
                        enable_working_hours=enable_work_hours,
                        auto_export_excel=auto_export_excel
                    )
                else:
                    print("已取消")

        except ValueError:
            print("❌ 输入无效，请输入数字")
    else:
        print("无效选择")


if __name__ == "__main__":
    main()
