#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
全量数据爬虫 - 所有批次 + 多种燃料类型 + Excel导出
目标: 爬取所有历史批次(1-403)的纯电动和混合动力车型
必选字段: 公告型号、批次、燃料类型、生产企业
"""

import urllib.request
import urllib.parse
import json
import time
import re
import os
from datetime import datetime
from typing import List, Dict, Optional
from http.cookiejar import CookieJar

# 燃料类型映射
FUEL_TYPES = {
    'C': '纯电动',      # rylx='C'
    'O': '混合动力',    # rylx='O' ✅ 已确认
}


class FullDataCrawler:
    """全量数据爬虫"""

    def __init__(self):
        self.base_url = "https://www.jdcsww.com"
        self.cookie_jar = CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar)
        )
        self.output_dir = 'data/full_data'
        os.makedirs(self.output_dir, exist_ok=True)

    def fetch_with_session(self, url: str, referer: str = None) -> Optional[str]:
        """使用会话获取页面"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Accept-Encoding': 'identity',
                'Connection': 'keep-alive',
            }

            if referer:
                headers['Referer'] = referer

            req = urllib.request.Request(url, headers=headers)
            with self.opener.open(req, timeout=30) as response:
                return response.read().decode('utf-8', errors='ignore')
        except Exception as e:
            print(f"      ✗ 获取失败: {e}")
            return None

    def parse_list_page_simple(self, html: str, batch: int, fuel_type_code: str) -> List[Dict]:
        """解析列表页 - 只提取必选字段"""
        vehicles = []

        # 提取车辆块
        vehicle_pattern = r'<li[^>]*>.*?<span class="car_name"><a href="([^"]+)"[^>]*>([^<]+)</a></span>.*?</li>'
        vehicle_matches = re.findall(vehicle_pattern, html, re.DOTALL)

        fuel_type_name = FUEL_TYPES.get(fuel_type_code, fuel_type_code)

        for idx, (detail_link, car_name) in enumerate(vehicle_matches, 1):
            # 基础信息
            vehicle = {
                '批次': batch,
                '燃料类型': fuel_type_name,
                '车辆名称': car_name.strip(),
                '详情链接': f"{self.base_url}{detail_link}",
            }

            # 提取公告编号
            bh_match = re.search(r'bh=([^&"\']+)', detail_link)
            if bh_match:
                vehicle['公告编号'] = bh_match.group(1)

            # 从列表页提取更多信息
            detail_link_escaped = re.escape(detail_link)
            vehicle_block_pattern = rf'<li[^>]*>.*?<a href="{detail_link_escaped}".*?</li>'
            vehicle_block_match = re.search(vehicle_block_pattern, html, re.DOTALL)

            if vehicle_block_match:
                block = vehicle_block_match.group(0)

                # 提取公告型号
                model_pattern = r'<span class="car_model">([^<]+?)</span>'
                model_match = re.search(model_pattern, block)
                if model_match:
                    vehicle['公告型号'] = model_match.group(1).strip()

                # 提取生产企业
                company_pattern = r'车辆生产企业名称：([^<\n]+?)(?:<|$)'
                company_match = re.search(company_pattern, block)
                if company_match:
                    vehicle['生产企业'] = company_match.group(1).strip()
                else:
                    # 尝试其他方式提取
                    label_pattern = r'<label>生产企业[：:]([^<]+?)</label>'
                    label_match = re.search(label_pattern, block)
                    if label_match:
                        vehicle['生产企业'] = label_match.group(1).strip()

            vehicles.append(vehicle)

        return vehicles

    def get_list_data(self, batch: int, fuel_type_code: str) -> List[Dict]:
        """获取单个批次的列表数据"""
        url = f"{self.base_url}/qcggs"
        params = {
            'ggxh': '',
            'ggpc': str(batch),
            'zwpp': '',
            'clmc': '1',
            'fdjxh': '',
            'qymc': '',
            'cph': '',
            'rylx': fuel_type_code,  # 燃料类型代码
            'viewtype': '0'
        }
        list_url = f"{url}?{urllib.parse.urlencode(params)}"

        print(f"  [{fuel_type_code}] 批次{batch}...")

        # 先访问列表页建立会话
        html = self.fetch_with_session(list_url)

        if not html:
            print(f"      ✗ 失败")
            return []

        # 检查是否被拦截
        if '非正常访问' in html or '爬取数据嫌疑' in html:
            print(f"      ✗ 被拦截")
            return []

        # 解析数据
        vehicles = self.parse_list_page_simple(html, batch, fuel_type_code)
        print(f"      ✓ 找到 {len(vehicles)} 辆")
        return vehicles

    def crawl_all_batches(self, start_batch: int = 1, end_batch: int = 403,
                          fuel_types: List[str] = ['C', 'H']) -> List[Dict]:
        """爬取所有批次"""
        print("=" * 70)
        print("全量数据爬虫")
        print("=" * 70)
        print(f"批次范围: {start_batch} - {end_batch}")
        print(f"燃料类型: {', '.join([FUEL_TYPES.get(ft, ft) for ft in fuel_types])}")
        print(f"预计总批次数: {(end_batch - start_batch + 1) * len(fuel_types)}")
        print("=" * 70)
        print()

        all_vehicles = []
        batch_count = 0
        error_batches = []

        # 进度跟踪
        total_tasks = (end_batch - start_batch + 1) * len(fuel_types)
        completed_tasks = 0

        for batch in range(start_batch, end_batch + 1):
            batch_has_data = False

            for fuel_type_code in fuel_types:
                batch_count += 1

                # 爬取数据
                vehicles = self.get_list_data(batch, fuel_type_code)

                if vehicles:
                    all_vehicles.extend(vehicles)
                    batch_has_data = True

                    # 显示进度
                    completed_tasks += 1
                    progress = (completed_tasks / total_tasks) * 100
                    print(f"  进度: {completed_tasks}/{total_tasks} ({progress:.1f}%) | 累计: {len(all_vehicles)} 辆")
                else:
                    error_batches.append(f"{batch}-{fuel_type_code}")

                # 延迟避免过快请求
                if batch < end_batch or fuel_type_code != fuel_types[-1]:
                    time.sleep(2)

            # 每完成一个批次，保存一次增量数据
            if batch_has_data:
                self.save_incremental(all_vehicles, batch)
                print()

        print("=" * 70)
        print(f"✅ 爬取完成！")
        print(f"✓ 总计: {len(all_vehicles)} 辆车")
        print(f"✓ 成功批次: {batch_count - len(error_batches)}")
        if error_batches:
            print(f"✗ 失败批次 ({len(error_batches)}): {', '.join(error_batches[:10])}...")
        print("=" * 70)

        return all_vehicles

    def save_incremental(self, vehicles: List[Dict], last_batch: int):
        """保存增量数据"""
        # 保存JSON备份
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        json_file = os.path.join(self.output_dir, f'incremental_batch{last_batch}_{timestamp}.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(vehicles, f, ensure_ascii=False, indent=2)
        print(f"  💾 增量保存: {os.path.basename(json_file)}")

    def export_to_excel(self, vehicles: List[Dict], filename: str = None):
        """导出到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("⚠️  需要安装 openpyxl: pip3 install openpyxl")
            return

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'vehicles_full_data_{timestamp}.xlsx'

        filepath = os.path.join(self.output_dir, filename)

        print(f"\n📊 正在导出Excel: {filename}")
        print(f"   总计: {len(vehicles)} 辆车")

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "车型数据"

        # 定义列
        columns = [
            ('批次', 10),
            ('燃料类型', 12),
            ('公告型号', 25),
            ('生产企业', 30),
            ('公告编号', 20),
            ('车辆名称', 40),
        ]

        # 设置列宽
        for col_idx, (col_name, width) in enumerate(columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, (col_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据样式
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # 写入数据
        for row_idx, vehicle in enumerate(vehicles, 2):
            for col_idx, (col_name, _) in enumerate(columns, 1):
                value = vehicle.get(col_name, '')

                # 处理None值
                if value is None:
                    value = ''

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

                # 冻结首行
                ws.freeze_panes = 'A2'

        # 保存
        wb.save(filepath)
        print(f"✓ Excel已保存: {filepath}")
        print(f"  文件大小: {os.path.getsize(filepath) / 1024 / 1024:.2f} MB")

        return filepath


def main():
    """主函数"""
    crawler = FullDataCrawler()

    # 先测试几个批次
    print("=" * 70)
    print("测试模式 - 爬取最近3个批次")
    print("=" * 70)

    # 测试: 爬取403, 402, 401批次
    vehicles = crawler.crawl_all_batches(
        start_batch=401,
        end_batch=403,
        fuel_types=['C']  # 先只爬纯电动
    )

    if vehicles:
        # 导出到Excel
        crawler.export_to_excel(vehicles, 'vehicles_test_401-403.xlsx')

        print("\n" + "=" * 70)
        print("📋 数据预览:")
        print("=" * 70)
        for v in vehicles[:5]:
            print(f"\n批次{v['批次']} | {v['燃料类型']}")
            print(f"  公告型号: {v.get('公告型号', 'N/A')}")
            print(f"  生产企业: {v.get('生产企业', 'N/A')}")
            print(f"  车辆名称: {v['车辆名称'][:40]}...")

        print("\n" + "=" * 70)
        print("✅ 测试完成！")
        print("\n如需爬取全部批次，请修改:")
        print("  crawler.crawl_all_batches(start_batch=1, end_batch=403, fuel_types=['C', 'H'])")
        print("=" * 70)


if __name__ == "__main__":
    main()
