#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生产环境全量爬虫
功能:
1. 爬取所有批次(1-403)的新能源车辆数据
2. 支持增量保存和断点续传
3. 导出为Excel格式
4. 必选字段: 公告型号、批次、燃料类型、生产企业
"""

import urllib.request
import urllib.parse
import json
import time
import re
import os
import logging
import random
from datetime import datetime
from typing import List, Dict, Optional, Set
from http.cookiejar import CookieJar


class ProductionCrawler:
    """生产环境爬虫"""

    def __init__(self):
        self.base_url = "https://www.jdcsww.com"
        self.cookie_jar = CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar)
        )
        self.output_dir = 'data/production'
        os.makedirs(self.output_dir, exist_ok=True)

        # 状态跟踪
        self.state_file = os.path.join(self.output_dir, 'crawler_state.json')
        self.data_file = os.path.join(self.output_dir, 'all_vehicles.json')
        self.excel_file = os.path.join(self.output_dir, 'vehicles_all_batches.xlsx')

        # 配置日志
        self._setup_logger()

    def _setup_logger(self):
        """配置日志系统"""
        # 创建日志记录器
        self.logger = logging.getLogger('ProductionCrawler')
        self.logger.setLevel(logging.INFO)

        # 清除现有处理器（避免重复）
        if self.logger.handlers:
            self.logger.handlers.clear()

        # 创建日志格式
        formatter = logging.Formatter(
            '%(asctime)s | %(levelname)-8s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

        # 文件处理器 - 保存到文件
        log_file = os.path.join(self.output_dir, f'crawler_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)

        # 控制台处理器 - 实时显示
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)

        # 添加处理器
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)

        # 保存日志文件路径
        self.log_file = log_file

        # 记录启动
        self.logger.info("=" * 70)
        self.logger.info("爬虫启动")
        self.logger.info(f"日志文件: {log_file}")
        self.logger.info("=" * 70)

    def load_state(self) -> Dict:
        """加载爬虫状态"""
        if os.path.exists(self.state_file):
            with open(self.state_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {
            'completed_batches': [],
            'failed_batches': [],
            'last_update': None,
            'total_vehicles': 0
        }

    def save_state(self, state: Dict):
        """保存爬虫状态"""
        state['last_update'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(self.state_file, 'w', encoding='utf-8') as f:
            json.dump(state, f, ensure_ascii=False, indent=2)

    def load_data(self) -> List[Dict]:
        """加载已有数据"""
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []

    def save_data(self, vehicles: List[Dict]):
        """保存所有数据"""
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(vehicles, f, ensure_ascii=False, indent=2)

    def fetch_with_session(self, url: str, referer: str = None, max_retries: int = 3) -> Optional[str]:
        """使用会话获取页面"""
        for attempt in range(max_retries):
            try:
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'zh-CN,zh;q=0.9',
                    'Accept-Encoding': 'identity',
                    'Connection': 'keep-alive',
                }

                if referer:
                    headers['Referer'] = referer

                req = urllib.request.Request(url, headers=headers)
                with self.opener.open(req, timeout=30) as response:
                    html = response.read().decode('utf-8', errors='ignore')

                    # 检查是否被拦截
                    if '非正常访问' in html or '爬取数据嫌疑' in html:
                        if attempt < max_retries - 1:
                            print(f"      ⚠️  被拦截，重试 {attempt + 1}/{max_retries}")
                            time.sleep(5)
                            continue
                        else:
                            return None

                    return html

            except Exception as e:
                if attempt < max_retries - 1:
                    retry_delay = 5 + (attempt * 5)  # 递增延迟: 5秒, 10秒, 15秒
                    self.logger.warning(f"网络错误，{retry_delay}秒后重试 {attempt + 1}/{max_retries}: {e}")
                    print(f"      ⚠️  网络错误，{retry_delay}秒后重试 {attempt + 1}/{max_retries}: {e}")
                    time.sleep(retry_delay)
                else:
                    self.logger.error(f"获取失败: {e}")
                    print(f"      ✗ 获取失败: {e}")
                    return None

        return None

    def parse_list_page(self, html: str, batch: int, fuel_type_name: str) -> List[Dict]:
        """解析列表页 - 提取必选字段"""
        vehicles = []

        # 提取车辆块
        vehicle_pattern = r'<li[^>]*>.*?<span class="car_name"><a href="([^"]+)"[^>]*>([^<]+)</a></span>.*?</li>'
        vehicle_matches = re.findall(vehicle_pattern, html, re.DOTALL)

        for idx, (detail_link, car_name) in enumerate(vehicle_matches, 1):
            vehicle = {
                '批次': batch,
                '燃料类型': fuel_type_name,
                '车辆名称': car_name.strip(),
                '详情链接': f"{self.base_url}{detail_link}",
            }

            # 提取公告编号
            bh_match = re.search(r'bh=([^&"\']+)', detail_link)
            if bh_match:
                vehicle['公告编号'] = bh_match.group(1)

            # 从列表页提取更多信息
            detail_link_escaped = re.escape(detail_link)
            vehicle_block_pattern = rf'<li[^>]*>.*?<a href="{detail_link_escaped}".*?</li>'
            vehicle_block_match = re.search(vehicle_block_pattern, html, re.DOTALL)

            if vehicle_block_match:
                block = vehicle_block_match.group(0)

                # 提取公告型号
                model_pattern = r'<span class="car_model">([^<]+?)</span>'
                model_match = re.search(model_pattern, block)
                if model_match:
                    vehicle['公告型号'] = model_match.group(1).strip()

                # 提取生产企业 - 多种模式
                company_patterns = [
                    r'车辆生产企业名称[：:]([^<\n]+?)(?:<|$)',
                    r'<label>生产企业[：:]([^<]+?)</label>',
                    r'<label>qymc[：:]([^<]+?)</label>',
                ]

                for pattern in company_patterns:
                    company_match = re.search(pattern, block)
                    if company_match:
                        vehicle['生产企业'] = company_match.group(1).strip()
                        break

                # 如果还没找到，尝试从label中提取
                if '生产企业' not in vehicle:
                    label_pattern = r'<label>([^<]*企业[^<]*)[：:]([^<]+?)</label>'
                    label_matches = re.findall(label_pattern, block)
                    for key, value in label_matches:
                        if '企业' in key:
                            vehicle['生产企业'] = value.strip()
                            break

            vehicles.append(vehicle)

        return vehicles

    def crawl_single_batch(self, batch: int, fuel_type_code: str, fuel_type_name: str) -> List[Dict]:
        """爬取单个批次"""
        url = f"{self.base_url}/qcggs"
        params = {
            'ggxh': '',
            'ggpc': str(batch),
            'zwpp': '',
            'clmc': '1',
            'fdjxh': '',
            'qymc': '',
            'cph': '',
            'rylx': fuel_type_code,
            'viewtype': '0'
        }
        list_url = f"{url}?{urllib.parse.urlencode(params)}"

        self.logger.info(f"开始爬取: 批次{batch} [{fuel_type_name}]")
        print(f"  批次{batch} [{fuel_type_name}]...")

        # 获取页面
        html = self.fetch_with_session(list_url)

        if not html:
            self.logger.warning(f"批次{batch} [{fuel_type_name}] - 获取页面失败")
            return []

        # 解析数据
        vehicles = self.parse_list_page(html, batch, fuel_type_name)
        self.logger.info(f"批次{batch} [{fuel_type_name}] - 完成，找到 {len(vehicles)} 辆")
        print(f"    ✓ {len(vehicles)} 辆")

        return vehicles

    def run_full_crawl(self, start_batch: int = 1, end_batch: int = 403):
        """执行全量爬取"""
        self.logger.info("=" * 70)
        self.logger.info(f"开始全量爬取: 批次 {start_batch}-{end_batch}")
        self.logger.info("=" * 70)

        print("=" * 70)
        print("生产环境全量爬虫")
        print("=" * 70)
        print(f"批次范围: {start_batch} - {end_batch}")
        print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"日志文件: {self.log_file}")
        print("=" * 70)
        print()

        # 加载状态
        state = self.load_state()
        all_vehicles = self.load_data()

        self.logger.info(f"已有数据: {len(all_vehicles)} 辆")
        self.logger.info(f"已完成批次: {len(state['completed_batches'])}")
        self.logger.info(f"失败批次: {len(state['failed_batches'])}")

        print(f"📊 已有数据: {len(all_vehicles)} 辆")
        print(f"✓ 已完成批次: {len(state['completed_batches'])}")
        print(f"✗ 失败批次: {len(state['failed_batches'])}")

        # 显示断点续传信息
        if state['completed_batches']:
            print(f"\n🔄 断点续传模式:")
            print(f"   将从已完成批次的下一个批次继续")
            if state['completed_batches']:
                last_completed = state['completed_batches'][-1]
                print(f"   上次完成: {last_completed}")
                batch_num = int(last_completed.split('-')[0])
                if batch_num < end_batch:
                    print(f"   下一个: 批次{batch_num + 1}")
        print()

        # 定义燃料类型
        fuel_types = [
            {'code': 'C', 'name': '纯电动'},
            {'code': 'O', 'name': '混合动力'},  # ✅ 已确认: rylx=O
        ]

        # 计算总任务数
        total_tasks = (end_batch - start_batch + 1) * len(fuel_types)
        completed_tasks = 0

        # 开始爬取
        for batch in range(start_batch, end_batch + 1):
            batch_vehicles = []

            for fuel_type in fuel_types:
                task_key = f"{batch}-{fuel_type['code']}"

                # ⭐ 跳过已完成的任务（断点续传核心）
                if task_key in state['completed_batches']:
                    completed_tasks += 1
                    self.logger.info(f"跳过已完成: {task_key}")
                    continue

                try:
                    # 爬取数据
                    vehicles = self.crawl_single_batch(batch, fuel_type['code'], fuel_type['name'])

                    if vehicles:
                        batch_vehicles.extend(vehicles)

                        # 更新状态
                        state['completed_batches'].append(task_key)
                        completed_tasks += 1

                        # 显示进度
                        progress = (completed_tasks / total_tasks) * 100
                        print(f"    进度: {completed_tasks}/{total_tasks} ({progress:.1f}%) | 累计: {len(all_vehicles) + len(batch_vehicles)} 辆")
                    else:
                        state['failed_batches'].append(task_key)

                    # 延迟 - 更大的随机范围，模拟真实用户行为
                    # 基础延迟: 7-12秒
                    # 随机波动: ±3秒
                    base_delay = random.uniform(7, 12)
                    random_jitter = random.uniform(-3, 3)
                    delay = max(5, base_delay + random_jitter)  # 最小5秒

                    self.logger.info(f"等待 {delay:.1f} 秒...")
                    time.sleep(delay)

                    # 定期休息 - 每50批次休息30秒
                    if completed_tasks > 0 and completed_tasks % 100 == 0:  # 50批次 × 2种燃料类型 = 100
                        rest_time = 30
                        self.logger.info(f"已爬取{completed_tasks}个任务，休息{rest_time}秒...")
                        print(f"\n  💦 已爬取{completed_tasks}个任务，休息{rest_time}秒...")
                        time.sleep(rest_time)
                        print(f"  ✓ 休息完成，继续爬取\n")

                except KeyboardInterrupt:
                    print("\n\n⚠️  用户中断！保存当前进度...")
                    all_vehicles.extend(batch_vehicles)
                    self.save_data(all_vehicles)
                    state['total_vehicles'] = len(all_vehicles)
                    self.save_state(state)
                    print("✓ 进度已保存")
                    return

                except Exception as e:
                    print(f"    ✗ 错误: {e}")
                    state['failed_batches'].append(task_key)

            # 合并数据
            if batch_vehicles:
                all_vehicles.extend(batch_vehicles)

                # 每完成一个批次保存一次
                self.save_data(all_vehicles)
                state['total_vehicles'] = len(all_vehicles)
                self.save_state(state)

                print(f"  💾 已保存 (累计 {len(all_vehicles)} 辆)")

            print()

        # 最终统计
        self.logger.info("=" * 70)
        self.logger.info("爬取完成")
        self.logger.info(f"总计: {len(all_vehicles)} 辆车")
        self.logger.info(f"成功: {len(state['completed_batches'])} 个任务")
        self.logger.info(f"失败: {len(state['failed_batches'])} 个任务")
        if state['failed_batches']:
            self.logger.warning(f"失败批次: {state['failed_batches'][:10]}")
        self.logger.info(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.logger.info("=" * 70)

        print("=" * 70)
        print("✅ 爬取完成！")
        print(f"✓ 总计: {len(all_vehicles)} 辆车")
        print(f"✓ 成功: {len(state['completed_batches'])} 个任务")
        print(f"✗ 失败: {len(state['failed_batches'])} 个任务")
        if state['failed_batches']:
            print(f"  失败列表: {state['failed_batches'][:10]}...")
        print(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"日志文件: {self.log_file}")
        print("=" * 70)
        print()

        # 导出Excel
        if all_vehicles:
            self.export_to_excel(all_vehicles)

        return all_vehicles

    def export_to_excel(self, vehicles: List[Dict]):
        """导出到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("⚠️  需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
            return

        filename = os.path.basename(self.excel_file)
        print(f"📊 正在导出Excel: {filename}")

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "车型数据"

        # 定义列
        columns = [
            ('批次', 10),
            ('燃料类型', 12),
            ('公告型号', 25),
            ('生产企业', 30),
            ('公告编号', 20),
            ('车辆名称', 40),
        ]

        # 设置列宽
        for col_idx, (col_name, width) in enumerate(columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 样式
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # 写入表头
        for col_idx, (col_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, vehicle in enumerate(vehicles, 2):
            for col_idx, (col_name, _) in enumerate(columns, 1):
                value = vehicle.get(col_name, '')
                if value is None:
                    value = ''

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存
        wb.save(self.excel_file)
        file_size = os.path.getsize(self.excel_file) / 1024 / 1024
        print(f"✓ Excel已保存: {self.excel_file}")
        print(f"  文件大小: {file_size:.2f} MB")
        print(f"  记录数: {len(vehicles)}")


def main():
    """主函数"""
    crawler = ProductionCrawler()

    print("\n" + "=" * 70)
    print("全量数据爬取")
    print("=" * 70)
    print("\n选项:")
    print("1. 爬取所有批次 (1-403)")
    print("2. 爬取指定范围")
    print("3. 继续上次中断的任务")
    print("4. 仅导出已有数据为Excel")
    print()

    choice = input("请选择 (1-4): ").strip()

    if choice == '1':
        crawler.run_full_crawl(start_batch=1, end_batch=403)
    elif choice == '2':
        start = int(input("起始批次: "))
        end = int(input("结束批次: "))
        crawler.run_full_crawl(start_batch=start, end_batch=end)
    elif choice == '3':
        state = crawler.load_state()
        all_vehicles = crawler.load_data()

        print("\n" + "=" * 70)
        print("📊 断点续传状态")
        print("=" * 70)
        print(f"已有数据: {len(all_vehicles)} 辆")
        print(f"已完成: {len(state['completed_batches'])} 个任务")

        if state['completed_batches']:
            print(f"\n✅ 已完成的批次 (最近5个):")
            for task in state['completed_batches'][-5:]:
                print(f"   - {task}")

        if state['failed_batches']:
            print(f"\n⚠️  失败批次: {len(state['failed_batches'])} 个")
            print(f"   {state['failed_batches'][:5]}...")
            retry = input("\n是否重试失败批次? (y/n): ").strip().lower()
            if retry == 'y':
                # 这里可以实现重试逻辑
                pass
        # 继续爬取
        crawler.run_full_crawl(start_batch=1, end_batch=403)
    elif choice == '4':
        all_vehicles = crawler.load_data()
        if all_vehicles:
            crawler.export_to_excel(all_vehicles)
        else:
            print("没有数据可导出")
    else:
        print("无效选择")


if __name__ == "__main__":
    main()
